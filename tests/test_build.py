import base64
import json
import re
import sys

import openpyxl
import pytest

import build

sys.path.insert(0, str(build.ROOT / "tools"))

import seed_excel  # noqa: E402

SKILLS_HEADER = ["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp"]
SKILLS_ROW = ["Hoch", "Ablenkung", "🎧", "Musik hören", "Ein Lied auflegen", "Kopfhörer"]


def js_funktion(quelltext: str, name: str) -> str:
    """Schneidet eine JavaScript-Funktion über passende Klammerzählung heraus.

    Verlässlicher als ein Regex, der am nächsten `}` am Zeilenanfang abbricht:
    das hinge an der Einrückung. Geschweifte Klammern in Zeichenketten oder
    Kommentaren zählt diese einfache Zählung allerdings mit — für die geprüften
    Funktionen kommen dort keine vor.
    """
    start = quelltext.index(f"function {name}(")
    tiefe = 0
    for i in range(quelltext.index("{", start), len(quelltext)):
        if quelltext[i] == "{":
            tiefe += 1
        elif quelltext[i] == "}":
            tiefe -= 1
            if tiefe == 0:
                return quelltext[start : i + 1]
    raise AssertionError(f"Funktion {name} wird nicht geschlossen")


def ohne_umbrueche(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def test_von_wird_gelesen_wenn_spalte_vorhanden(mappe, monkeypatch):
    pfad = mappe(SKILLS_HEADER + ["Von"], [SKILLS_ROW + ["Max"]])
    monkeypatch.setattr(build, "XLSX", pfad)
    daten = build.load_data()
    assert daten["hoch"]["kategorien"][0]["skills"][0]["von"] == "Max"


def test_von_ist_leer_wenn_spalte_fehlt(mappe, monkeypatch):
    pfad = mappe(SKILLS_HEADER, [SKILLS_ROW])
    monkeypatch.setattr(build, "XLSX", pfad)
    daten = build.load_data()
    assert daten["hoch"]["kategorien"][0]["skills"][0]["von"] == ""


def test_pflichtspalte_fehlt_bricht_ab(mappe, monkeypatch):
    pfad = mappe(["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung"], [SKILLS_ROW[:5]])
    monkeypatch.setattr(build, "XLSX", pfad)
    try:
        build.load_data()
    except build.BuildError as fehler:
        assert "Tipp" in str(fehler)
    else:
        raise AssertionError("BuildError erwartet")


def test_vorlage_enthaelt_namenszeile():
    vorlage = build.TEMPLATE.read_bytes().decode("utf-8-sig")
    assert 'id="m-von"' in vorlage
    assert "modal-von" in vorlage

    # Nicht nur "s.von"/"s.erg" irgendwo suchen (koennte ein Kommentar sein),
    # sondern den zusammenhaengenden Anzeige-Block aus openModal pruefen: er
    # muss die Namenszeile bei vorhandenem "von" und/oder "erg" zusammensetzen
    # UND bei beiden fehlend wieder verstecken/leeren (sonst bliebe der Name
    # eines fruehreren Skills stehen).
    treffer = re.search(r"var teile=\[\];.*?else\s*\{.*?\}", vorlage, re.DOTALL)
    assert treffer, "Anzeige-Block fuer s.von/s.erg in openModal nicht gefunden"
    block = re.sub(r"\s+", " ", treffer.group(0)).strip()
    assert block == (
        "var teile=[]; if(s.von){ teile.push('Vorgeschlagen von '+s.von); } "
        "if(s.erg){ teile.push('Ergänzt von '+s.erg); } "
        "if(teile.length){ mVon.textContent=teile.join(' · '); mVon.hidden=false; } "
        "else { mVon.textContent=''; mVon.hidden=true; }"
    )


def test_daten_json_wird_geschrieben(mappe, monkeypatch, tmp_path):
    pfad = mappe(SKILLS_HEADER + ["Von"], [SKILLS_ROW + ["Max"]])
    ziel = tmp_path / "skills-daten.json"
    monkeypatch.setattr(build, "XLSX", pfad)
    monkeypatch.setattr(build, "DATEN_JSON", ziel)

    build.write_daten_json(build.load_data())

    roh = ziel.read_bytes()
    assert not roh.startswith(b"\xef\xbb\xbf"), "JSON darf kein BOM haben"
    daten = json.loads(roh.decode("utf-8"))
    assert set(daten) == {"hoch", "mittel", "tief"}
    skill = daten["hoch"]["kategorien"][0]["skills"][0]
    assert skill["t"] == "Musik hören"
    assert skill["von"] == "Max"


def test_vorschlagsseite_wird_erzeugt(mappe, monkeypatch, tmp_path):
    pfad = mappe(SKILLS_HEADER + ["Von"], [SKILLS_ROW + ["Max"]])
    ziel = tmp_path / "skill-vorschlagen.html"
    monkeypatch.setattr(build, "XLSX", pfad)
    monkeypatch.setattr(build, "OUTPUT_VORSCHLAG", ziel)

    build.render_vorschlag(build.load_data())

    roh = ziel.read_bytes()
    assert roh.startswith(b"\xef\xbb\xbf")
    html = roh.decode("utf-8-sig")
    assert "var DATEN = {" in html
    assert "Ablenkung" in html
    assert "/*__BUILD_DATA__*/" not in html


def test_script_block_kann_nicht_verlassen_werden(mappe, monkeypatch, tmp_path):
    """Ein </script> im Freitext darf das Skriptelement nicht beenden."""
    gift = "</script><script>alert(1)</script>"
    zeile = SKILLS_ROW[:4] + [gift, "Kopfhörer"]
    pfad = mappe(SKILLS_HEADER, [zeile])
    ziel = tmp_path / "skillsliste.html"
    monkeypatch.setattr(build, "XLSX", pfad)
    monkeypatch.setattr(build, "OUTPUT", ziel)

    build.render(build.load_data(), {})

    html = ziel.read_bytes().decode("utf-8-sig")
    assert gift not in html, "die Zeichenfolge darf nicht wörtlich im Erzeugnis stehen"
    assert "\\u003c/script\\u003e" in html
    assert "<script>alert(1)" not in html

    # Der Wert selbst bleibt unverändert – der Browser (bzw. json.loads) setzt
    # die Escapes beim Parsen zurück.
    # Die Datenzeile ist eine einzige Zeile: bis zum Zeilenende lesen, ohne ";".
    roh = html.split("var DATA = ", 1)[1].split("\n", 1)[0].rstrip().rstrip(";")
    wieder = json.loads(roh)
    assert wieder["hoch"]["kategorien"][0]["skills"][0]["b"] == gift


def test_daten_json_endet_mit_zeilenumbruch(mappe, monkeypatch, tmp_path):
    """Feste Zeilenenden, sonst gibt es plattformabhängige Scheinunterschiede."""
    pfad = mappe(SKILLS_HEADER, [SKILLS_ROW])
    ziel = tmp_path / "skills-daten.json"
    monkeypatch.setattr(build, "XLSX", pfad)
    monkeypatch.setattr(build, "DATEN_JSON", ziel)

    build.write_daten_json(build.load_data())

    roh = ziel.read_bytes()
    assert roh.endswith(b"\n")
    assert b"\r\n" not in roh


def test_vorschlagsvorlage_hat_pflichtbestandteile():
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    for baustein in ['name="falle"', "cf-turnstile", "footer-credit", "WORKER_URL"]:
        assert baustein in vorlage, baustein


def test_vorschlagsvorlage_hat_keine_platzhalter_mehr():
    """Fängt ab, dass die Seite mit unersetzter Worker-Adresse veröffentlicht wird."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    for platzhalter in ("WORKER_URL_HIER_EINSETZEN", "TURNSTILE_SITEKEY_HIER_EINSETZEN"):
        assert platzhalter not in vorlage, platzhalter
    assert "https://" in vorlage.split('var WORKER_URL = "')[1][:60]


def test_vorschlagsvorlage_setzt_die_turnstile_aktion():
    """Ohne data-action lehnt der Worker jede Einreichung ab."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    assert 'data-action="skill-vorschlag"' in vorlage


def test_vorschlagsvorlage_emojifeld_hat_kein_maxlength_und_bietet_auswahl():
    """maxlength="2" zaehlt UTF-16-Einheiten und blockiert zusammengesetzte
    Emoji (z. B. 🧘‍♀️) stumm. Stattdessen muss die Vorlage eine eingebettete
    Emoji-Auswahl anbieten (kein Nachladen, keine fremde Bibliothek)."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    emoji_feld = re.search(r'<input[^>]*id="emoji"[^>]*>', vorlage)
    assert emoji_feld, "Emoji-Eingabefeld nicht gefunden"
    assert "maxlength" not in emoji_feld.group(0)
    assert "Intl.Segmenter" in vorlage
    assert "Schon verwendet" in vorlage


def test_vorschlagsvorlage_emoji_pruefung_hat_notbremse_und_rueckfallweg():
    """Die Client-Pruefung muss dieselbe 16-Codepunkt-Notbremse haben wie
    worker/validate.js – sonst weist der Worker eine sehr lange, aber als EIN
    Graphem zaehlende ZWJ-Kette ab, obwohl das Formular sie durchliess (die
    Person sieht dann eine unerklaerliche Meldung nach dem Absenden). Fehlt
    Intl.Segmenter (aeltere Firefox-Fassungen), darf das Absenden nicht mit
    einem unbehandelten Fehler abbrechen."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    assert "EMOJI_CODEPUNKT_GRENZE" in vorlage
    assert "typeof Intl.Segmenter" in vorlage
    assert "Array.from(wert).length" in vorlage


def test_vorschlagsvorlage_themengruppen_stehen_vor_schon_verwendet():
    """'Schon verwendet' kann viele Eintraege haben und darf die kuratierten
    Themengruppen darum nicht unter die Scroll-Kante schieben."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    erste_themengruppe = vorlage.index("Körper & Bewegung")
    schon_verwendet_literal = vorlage.index("'Schon verwendet'")
    assert erste_themengruppe < schon_verwendet_literal


def test_erg_wird_gelesen_wenn_spalte_vorhanden(mappe, monkeypatch):
    pfad = mappe(SKILLS_HEADER + ["Von", "Ergaenzt"], [SKILLS_ROW + ["Max", "Lea"]])
    monkeypatch.setattr(build, "XLSX", pfad)
    daten = build.load_data()
    skill = daten["hoch"]["kategorien"][0]["skills"][0]
    assert skill["von"] == "Max"
    assert skill["erg"] == "Lea"


def test_erg_ist_leer_wenn_spalte_fehlt(mappe, monkeypatch):
    pfad = mappe(SKILLS_HEADER, [SKILLS_ROW])
    monkeypatch.setattr(build, "XLSX", pfad)
    skill = build.load_data()["hoch"]["kategorien"][0]["skills"][0]
    assert skill["erg"] == ""


def test_vorlage_nennt_beide_beitragenden():
    vorlage = build.TEMPLATE.read_bytes().decode("utf-8-sig")
    assert 'id="m-von"' in vorlage
    assert "s.erg" in vorlage
    assert "Ergänzt von" in vorlage


def test_vorschlagsvorlage_hat_zwei_reiter():
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    assert 'id="reiter-neu"' in vorlage
    assert 'id="reiter-aenderung"' in vorlage
    assert 'role="tab"' in vorlage


def test_vorschlagsvorlage_sendet_die_aenderungsart():
    """Ohne art und original kann der Worker eine Aenderung nicht zuordnen."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    # Einfache Anführungszeichen: die Vorlage schreibt JavaScript durchgängig so.
    assert "art:'aenderung'" in vorlage
    assert "original:el('original').value" in vorlage

    # Der bestehende Weg bleibt unverändert. Nicht auf "art:" prüfen — das wäre
    # ein Teilwort und ginge an einem künftigen "start:" grundlos kaputt.
    # Stattdessen die Schlüssel des Zweigs vollständig festnageln.
    neuer_zweig = vorlage.split("} : {", 1)[1].split("})", 1)[0]
    assert re.findall(r"^\s*(\w+):", neuer_zweig, re.M) == [
        "stufe", "kategorie", "emoji", "titel", "beschreibung", "tipp",
        "von", "links", "falle", "turnstile",
    ]


def test_vorschlagsvorlage_bewahrt_den_entwurf_beim_reiterwechsel():
    """Ein schon getippter Text darf durch einen Blick in den anderen Reiter
    nicht verloren gehen — wer etwas Persönliches beschrieben hat, tippt es kein
    zweites Mal. Zugleich darf der vorausgefüllte Text einer Änderung NIE in den
    Entwurf für einen neuen Skill geraten, sonst wird ein bestehender Skill als
    neuer eingereicht. Beides hängt daran, dass gesichert wird, solange `modus`
    noch der alte ist."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    assert "var ENTWURF_FELDER=['emoji','titel','beschreibung','tipp','von'];" in vorlage

    rumpf = ohne_umbrueche(js_funktion(vorlage, "modusSetzen"))
    # Ein Klick auf den ohnehin aktiven Reiter (oder Pos1/Ende) darf gar nichts
    # anfassen — sonst füllt er den schon bearbeiteten Text neu vor.
    assert rumpf.index("if(neuerModus===modus){ return; }") < rumpf.index("entwurfSichern();")
    assert rumpf.index("entwurfSichern();") < rumpf.index("modus=neuerModus;")
    assert "felderSetzen(entwuerfe.neu);" in rumpf
    assert "value=''" not in rumpf, "beim Rückwechsel wird zurückgeschrieben, nicht geleert"


def test_vorschlagsvorlage_haelt_beide_entwuerfe_getrennt():
    """Beide Reiter haben einen Entwurf, und sie dürfen sich nicht vermischen.
    Zum Änderungs-Entwurf gehört die getroffene Skill-Auswahl dazu: wer
    'Scharfer Senf' gewählt und bearbeitet hat, muss beides zurückbekommen und
    nicht wieder den ersten Eintrag der Liste."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    assert "var entwuerfe={neu:null,aenderung:null};" in vorlage

    sichern = ohne_umbrueche(js_funktion(vorlage, "entwurfSichern"))
    # Unter dem gerade aktiven Modus ablegen, nie unter einem festen Schlüssel.
    assert "entwuerfe[modus]=e;" in sichern
    assert "if(modus==='aenderung'){ e.original=el('original').value; }" in sichern

    betreten = ohne_umbrueche(js_funktion(vorlage, "aenderungBetreten"))
    assert "el('original').value=e.original;" in betreten
    # Stufe/Kategorie können sich zwischendurch geändert haben: eine Auswahl,
    # die es nicht mehr gibt, darf nicht zurückgeschrieben werden.
    assert "skillsDerKategorie()" in betreten
    assert "entwuerfe.neu" not in betreten, "die beiden Puffer bleiben getrennt"


def test_vorschlagsvorlage_holt_den_namen_auch_beim_rueckfall_zurueck():
    """Gibt es den gemerkten Skill in der eingestellten Stufe/Kategorie nicht
    mehr, bleiben die Skillfelder frisch vorausgefüllt — der Name gehört aber der
    Person und nicht dem Skill. Käme er nur auf dem Normalweg zurück, stünde im
    Änderungs-Reiter plötzlich der Name aus dem Neu-Reiter, und abgeschickt würde
    ein Name, den die Person in diesem Reiter nie eingetragen hat."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    betreten = ohne_umbrueche(js_funktion(vorlage, "aenderungBetreten"))
    assert betreten.index("feldSetzen('von', e.von);") < betreten.index(
        "if(!vorhanden){ return; }"
    ), "der Name muss zurückkommen, bevor der Rückfall aussteigt"


def test_vorschlagsvorlage_reiterleiste_ist_vollstaendig_ausgezeichnet():
    """role=tab ohne zugehörigen tabpanel kündigt einem Screenreader einen
    Bereich an, den es nicht gibt. Und ohne Pfeiltasten samt rollendem tabindex
    ist eine Reiterleiste mit der Tastatur nicht wie erwartet bedienbar."""
    vorlage = build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")
    assert 'role="tabpanel"' in vorlage
    assert 'aria-labelledby="reiter-neu"' in vorlage
    assert "el('formular').setAttribute('aria-labelledby'" in vorlage
    assert "el('reiter-neu').tabIndex=" in vorlage
    assert "el('reiter-aenderung').tabIndex=" in vorlage
    assert "ArrowLeft" in vorlage and "ArrowRight" in vorlage

    # Nach dem Absenden verschwindet das Formular. Bleibt die Leiste stehen,
    # zeigen zwei Reiter auf einen unsichtbaren Bereich und tun sichtbar nichts.
    assert "el('reiterleiste').hidden=true;" in vorlage
    # display:flex schlägt das [hidden] des Browsers — ohne diese Regel bliebe
    # die Leiste trotz hidden sichtbar.
    assert ".reiter[hidden]{display:none}" in vorlage


def test_skills_werden_innerhalb_der_kategorie_alphabetisch_sortiert(mappe, monkeypatch):
    """Damit die Liste unabhaengig von der Zeilenfolge in der Excel gleich aussieht."""
    pfad = mappe(
        SKILLS_HEADER,
        [
            ["Hoch", "Ablenkung", "🎧", "Zitrone", "Sauer.", ""],
            ["Hoch", "Ablenkung", "🎵", "Ammoniak", "Scharf.", ""],
            ["Hoch", "Ablenkung", "🌶️", "Musik hören", "Laut.", ""],
        ],
    )
    monkeypatch.setattr(build, "XLSX", pfad)
    daten = build.load_data()
    titel = [s["t"] for s in daten["hoch"]["kategorien"][0]["skills"]]
    assert titel == ["Ammoniak", "Musik hören", "Zitrone"]


def test_umlaute_sortieren_wie_im_woerterbuch(mappe, monkeypatch):
    """Ohne Sonderbehandlung stuende 'Duftoele' hinter 'Duftzone' – das oe liegt
    im Zeichensatz hinter dem z."""
    pfad = mappe(
        SKILLS_HEADER,
        [
            ["Hoch", "Ablenkung", "🎧", "Duftzone", "Text.", ""],
            ["Hoch", "Ablenkung", "🎵", "Duftöle", "Text.", ""],
            ["Hoch", "Ablenkung", "🌶️", "Duftbaum", "Text.", ""],
        ],
    )
    monkeypatch.setattr(build, "XLSX", pfad)
    daten = build.load_data()
    titel = [s["t"] for s in daten["hoch"]["kategorien"][0]["skills"]]
    assert titel == ["Duftbaum", "Duftöle", "Duftzone"]


def test_sortierung_haengt_nicht_an_der_zeilenfolge(mappe, monkeypatch):
    """Dieselben Skills in umgekehrter Zeilenfolge ergeben dieselbe Liste."""
    zeilen = [
        ["Hoch", "Ablenkung", "🎧", "Aal", "Text.", ""],
        ["Hoch", "Ablenkung", "🎵", "Ärger", "Text.", ""],
        ["Hoch", "Ablenkung", "🌶️", "Aber", "Text.", ""],
    ]
    monkeypatch.setattr(build, "XLSX", mappe(SKILLS_HEADER, zeilen))
    vorwaerts = [s["t"] for s in build.load_data()["hoch"]["kategorien"][0]["skills"]]
    monkeypatch.setattr(build, "XLSX", mappe(SKILLS_HEADER, list(reversed(zeilen))))
    rueckwaerts = [s["t"] for s in build.load_data()["hoch"]["kategorien"][0]["skills"]]
    assert vorwaerts == rueckwaerts
    assert vorwaerts == ["Aal", "Aber", "Ärger"]


def test_kategorien_behalten_ihre_reihenfolge(mappe, monkeypatch):
    """Sortiert wird NUR innerhalb einer Kategorie – die Gruppierung bleibt."""
    pfad = mappe(
        SKILLS_HEADER,
        [
            ["Hoch", "Ablenkung", "🎧", "Zitrone", "Text.", ""],
            ["Hoch", "Anti-Craving", "🎵", "Ammoniak", "Text.", ""],
        ],
        kategorien_rows=[["Hoch", "Ablenkung", "🎧"], ["Hoch", "Anti-Craving", "🌶️"]],
    )
    monkeypatch.setattr(build, "XLSX", pfad)
    kategorien = build.load_data()["hoch"]["kategorien"]
    assert [k["label"] for k in kategorien] == ["Ablenkung", "Anti-Craving"]
    assert [s["t"] for s in kategorien[0]["skills"]] == ["Zitrone"]


LINK_HEADER = SKILLS_HEADER + ["Link1", "Link2", "Link3"]


def erste_skills(daten):
    return daten["hoch"]["kategorien"][0]["skills"][0]


def test_links_werden_gelesen(mappe, monkeypatch):
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://www.skillsbox.ch/p/1", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == [
        {"u": "https://www.skillsbox.ch/p/1", "t": ""}
    ]


def test_links_sind_leer_wenn_spalten_fehlen(mappe, monkeypatch):
    pfad = mappe(SKILLS_HEADER, [SKILLS_ROW])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == []


def test_luecke_wird_zusammengeschoben(mappe, monkeypatch):
    """Wer den ersten von zwei Links entfernt, soll die uebrigen nicht
    von Hand aufruecken muessen."""
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["", "https://a.ch/x", "https://b.ch/y"]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == [
        {"u": "https://a.ch/x", "t": ""},
        {"u": "https://b.ch/y", "t": ""},
    ]


def test_doppelte_url_faellt_weg(mappe, monkeypatch):
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "https://a.ch/x", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == [{"u": "https://a.ch/x", "t": ""}]


@pytest.mark.parametrize(
    "url, teil",
    [
        ("http://a.ch/x", "https://"),
        ("https://bit.ly/abc", "Linkverkürzer"),
        ("https://192.168.0.1/x", "IP-Adresse"),
        ("https://a.ch:8080/x", "Portnummer"),
        ("https://wer:was@a.ch/x", "Benutzerangabe"),
        ("https://ohnepunkt/x", "Hostnamen"),
        ("https://a.ch/<script>", "Klammern"),
        ("ftp://a.ch/x", "https://"),
    ],
)
def test_ungueltiger_link_bricht_ab(mappe, monkeypatch, url, teil):
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + [url, "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    with pytest.raises(build.BuildError) as fehler:
        build.load_data()
    meldung = str(fehler.value)
    assert teil in meldung
    # Die Meldung muss Zeile UND Spalte nennen, sonst sucht man in der Excel.
    assert "Zeile 2" in meldung
    assert "Link1" in meldung


def test_zu_langer_link_bricht_ab(mappe, monkeypatch):
    lang = "https://a.ch/" + "x" * 300
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + [lang, "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    with pytest.raises(build.BuildError) as fehler:
        build.load_data()
    assert "zu lang" in str(fehler.value).lower()


def test_links_stehen_in_der_json(mappe, monkeypatch, tmp_path):
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    ziel = tmp_path / "skills-daten.json"
    monkeypatch.setattr(build, "DATEN_JSON", ziel)
    build.write_daten_json(build.load_data())
    daten = json.loads(ziel.read_text(encoding="utf-8"))
    assert daten["hoch"]["kategorien"][0]["skills"][0]["links"] == [
        {"u": "https://a.ch/x", "t": ""}
    ]


def test_vorlage_zeigt_bezugsquellen():
    vorlage = build.TEMPLATE.read_bytes().decode("utf-8-sig")
    assert 'id="m-links"' in vorlage
    assert 'id="m-links-liste"' in vorlage
    assert "Von Besuchern vorgeschlagen · keine Empfehlung, keine Provision" in vorlage

    block = ohne_umbrueche(js_funktion(vorlage, "openModal"))
    # Ohne Links bleibt der ganze Bereich verborgen – keine leere Überschrift.
    assert "mLinks.hidden = !(s.links && s.links.length);" in block
    # Ohne nofollow/ugc wäre die Seite ein lohnendes Ziel für Link-Spam.
    assert "a.rel='noopener noreferrer nofollow ugc';" in block
    # Der Bereich muss bei jedem Öffnen geleert werden, sonst stehen die
    # Quellen des zuvor angesehenen Skills noch da.
    assert "mLinksListe.textContent='';" in block


def test_gastgeber_kuerzt_www():
    vorlage = build.TEMPLATE.read_bytes().decode("utf-8-sig")
    rumpf = ohne_umbrueche(js_funktion(vorlage, "gastgeber"))
    # Exakte Fassung statt nur "catch" in rumpf: sonst koennte der catch-Rumpf
    # klanglos leer werden (z. B. "return '';") und der Dialog zeigte leere
    # Knöpfe, ohne dass dieser Test es merkt.
    assert rumpf == (
        "function gastgeber(u){ "
        "try{ return new URL(u).hostname.replace(/^www\\./,''); } "
        "catch(e){ return u; } "
        "}"
    )


def vorschlagsvorlage():
    return build.TEMPLATE_VORSCHLAG.read_bytes().decode("utf-8-sig")


def test_formular_hat_dynamische_linkliste():
    vorlage = vorschlagsvorlage()
    assert 'id="links-liste"' in vorlage
    assert 'id="link-plus"' in vorlage
    assert "var MAX_LINKS = 3;" in vorlage
    # type=url braucht dieselbe Gestaltung wie die uebrigen Felder, sonst
    # steht ein nacktes Eingabefeld mitten im Formular.
    assert "input[type=text],input[type=url],textarea,select{" in vorlage


def test_letzte_linkzeile_wird_geleert_statt_entfernt():
    vorlage = vorschlagsvorlage()
    rumpf = ohne_umbrueche(js_funktion(vorlage, "linkZeileBauen"))
    assert "if(el('links-liste').children.length>1){ zeile.remove(); }" in rumpf
    assert "else { feld.value=''; }" in rumpf


def test_entwurf_traegt_die_links_mit():
    vorlage = vorschlagsvorlage()
    # Ohne diese beiden Zeilen verlieren die Links beim Reiterwechsel ihren
    # Inhalt – genau der Fehler, gegen den es die Entwuerfe ueberhaupt gibt.
    assert "e.links=linksLesen();" in ohne_umbrueche(js_funktion(vorlage, "entwurfSichern"))
    assert "linksSetzen(e?e.links:[]);" in ohne_umbrueche(js_funktion(vorlage, "felderSetzen"))


def test_ergaenzen_fuellt_die_bestehenden_links_vor():
    vorlage = vorschlagsvorlage()
    rumpf = ohne_umbrueche(js_funktion(vorlage, "originalUebernehmen"))
    # Traegt das Vorausfuellen die Links nicht mit, wuerde eine Ergaenzung sie
    # loeschen: das Uebernahme-Skript ersetzt die Spalten vollstaendig. Das
    # Formular schickt weiterhin nur Adressen (l.u), die Beschriftung pflegt
    # allein die betreuende Person in der Excel.
    assert "linksSetzen((s.links||[]).map(function(l){ return l.u||l; }));" in rumpf


def test_beide_absende_ruempfe_schicken_links():
    assert vorschlagsvorlage().count("links:linksLesen(),") == 2


def test_seed_excel_kennt_die_linkspalten():
    """Ein Zurücksetzen der Mappe darf die Bezugsquellen nicht verlieren."""
    quelle = (build.ROOT / "tools" / "seed_excel.py").read_text(encoding="utf-8")
    assert '"Link1", "Link2", "Link3"' in quelle
    assert 's.get("links"' in quelle


def test_seed_excel_schreibt_die_links_richtig(monkeypatch, tmp_path):
    """Fuehrt seed_excel.main() wirklich aus (statt nur den Quelltext zu
    grep-en) und prueft die geschriebenen Zellen fuer 0, 1, 2 und 3 Links.
    Deckt genau das ab, was der Quelltext-Test oben nicht sehen wuerde: eine
    falsche Slice-Grenze, eine falsch herum aufgefuellte Liste oder vertauschte
    Spalten wuerden hier auffliegen, dort nicht.
    """
    synthetisch = {
        "hoch": {
            "kategorien": [
                {
                    "label": "Testkategorie",
                    "skills": [
                        {"e": "🎧", "t": "Ohne Link", "b": "b", "links": []},
                        {"e": "🎧", "t": "Ein Link", "b": "b",
                         "links": ["https://a.ch"]},
                        {"e": "🎧", "t": "Zwei Links", "b": "b",
                         "links": ["https://a.ch", "https://b.ch"]},
                        {"e": "🎧", "t": "Drei Links", "b": "b",
                         "links": ["https://a.ch", "https://b.ch", "https://c.ch"]},
                    ],
                }
            ]
        }
    }

    # Beweis, dass die Umleitung wirklich greift, bevor main() ueberhaupt
    # laeuft: main() liest OUT als Modul-Global bei jedem Aufruf frisch, ein
    # monkeypatch.setattr auf das Modulattribut wirkt darum zuverlaessig (im
    # Unterschied zu einem Default-Argument, das beim Definieren der Funktion
    # gebunden wuerde). Die echte Mappe bleibt Beweis-Byte-fuer-Byte unberuehrt.
    echte_mappe = seed_excel.ROOT / "skills_daten.xlsx"
    vorher = echte_mappe.read_bytes()
    ziel = tmp_path / "skills_daten.xlsx"
    monkeypatch.setattr(seed_excel, "load_data", lambda: synthetisch)
    monkeypatch.setattr(seed_excel, "OUT", ziel)

    seed_excel.main()

    assert echte_mappe.read_bytes() == vorher, "seed_excel hat die echte Mappe angefasst"
    assert ziel.exists()

    wb = openpyxl.load_workbook(ziel)
    ws = wb["Skills"]
    header = [c.value for c in ws[1]]
    assert header == ["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp",
                       "Von", "Ergaenzt", "Link1", "Link2", "Link3"]

    zeilen = {row[3].value: row for row in ws.iter_rows(min_row=2)}  # Titel -> Zeile

    # openpyxl schreibt eine leere Zeichenkette nicht als "" in die Zelle,
    # sondern laesst sie leer: beim erneuten Oeffnen liest eine so gespeicherte
    # Zelle als None zurueck, nicht als "". Darum wird unten gegen None
    # geprueft, nicht gegen "".
    def links_spalten(titel):
        zeile = zeilen[titel]
        return [zeile[8].value, zeile[9].value, zeile[10].value]  # Link1..3

    assert links_spalten("Ohne Link") == [None, None, None]
    assert links_spalten("Ein Link") == ["https://a.ch", None, None]
    assert links_spalten("Zwei Links") == ["https://a.ch", "https://b.ch", None]
    assert links_spalten("Drei Links") == ["https://a.ch", "https://b.ch", "https://c.ch"]


TEXT_HEADER = SKILLS_HEADER + ["Link1", "Text1", "Link2", "Text2", "Link3", "Text3"]


def test_beschriftung_wird_gelesen(mappe, monkeypatch):
    pfad = mappe(TEXT_HEADER, [SKILLS_ROW + ["https://a.ch/x", "Igelball", "", "", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == [
        {"u": "https://a.ch/x", "t": "Igelball"}
    ]


def test_ohne_beschriftung_bleibt_t_leer(mappe, monkeypatch):
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == [{"u": "https://a.ch/x", "t": ""}]


def test_beschriftung_ohne_adresse_bricht_ab(mappe, monkeypatch):
    """Sonst staende eine Beschriftung ohne Ziel in der Mappe und niemand saehe es."""
    pfad = mappe(TEXT_HEADER, [SKILLS_ROW + ["", "Igelball", "", "", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    with pytest.raises(build.BuildError) as fehler:
        build.load_data()
    meldung = str(fehler.value)
    assert "Text1" in meldung and "Link1" in meldung
    assert "Zeile 2" in meldung


@pytest.mark.parametrize(
    "text, teil",
    [
        ("x" * 31, "zu lang"),
        ("<b>Igelball", "Klammern"),
        ("Siehe http://a.ch", "Adresse"),
    ],
)
def test_ungueltige_beschriftung_bricht_ab(mappe, monkeypatch, text, teil):
    pfad = mappe(TEXT_HEADER, [SKILLS_ROW + ["https://a.ch/x", text, "", "", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    with pytest.raises(build.BuildError) as fehler:
        build.load_data()
    assert teil in str(fehler.value)
    assert "Text1" in str(fehler.value)


def test_genau_dreissig_zeichen_sind_erlaubt(mappe, monkeypatch):
    pfad = mappe(TEXT_HEADER, [SKILLS_ROW + ["https://a.ch/x", "x" * 30, "", "", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"][0]["t"] == "x" * 30


def test_beschriftungen_folgen_ihrem_link_beim_zusammenschieben(mappe, monkeypatch):
    """Luecke in Link1: Link2/Text2 ruecken gemeinsam auf, das Paar darf nicht
    auseinanderfallen."""
    pfad = mappe(
        TEXT_HEADER,
        [SKILLS_ROW + ["", "", "https://b.ch/y", "Massage", "https://c.ch/z", "Kette"]],
    )
    monkeypatch.setattr(build, "XLSX", pfad)
    assert erste_skills(build.load_data())["links"] == [
        {"u": "https://b.ch/y", "t": "Massage"},
        {"u": "https://c.ch/z", "t": "Kette"},
    ]


def test_favicon_wird_eingebettet(mappe, monkeypatch, tmp_path):
    ordner = tmp_path / "favicons"
    ordner.mkdir()
    # Kleinstes gueltiges PNG (1x1, transparent) - als base64, damit man es
    # nicht Byte fuer Byte pruefen muss.
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk"
        "YPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
    )
    (ordner / "a.ch.png").write_bytes(png)
    monkeypatch.setattr(build, "FAVICON_DIR", ordner)
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    icons = build.lade_favicons(build.load_data())
    assert list(icons) == ["a.ch"]
    assert icons["a.ch"].startswith("data:image/png;base64,")


def test_fehlendes_favicon_bricht_den_build_nicht(mappe, monkeypatch, tmp_path):
    """Wer eine Bezugsquelle bei einem neuen Haendler eintraegt, darf damit
    nicht die ganze Website blockieren."""
    leer = tmp_path / "leer"
    leer.mkdir()
    monkeypatch.setattr(build, "FAVICON_DIR", leer)
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert build.lade_favicons(build.load_data()) == {}


def test_gastgeber_entfernt_www():
    assert build.gastgeber("https://www.skills-box.ch/products/x") == "skills-box.ch"
    assert build.gastgeber("kein link") == ""


def test_vorlage_hat_den_icons_platzhalter():
    vorlage = build.TEMPLATE.read_bytes().decode("utf-8-sig")
    assert build.PLACEHOLDER_ICONS in vorlage


def test_unlesbares_favicon_bricht_den_build_nicht(mappe, monkeypatch, tmp_path):
    """Vorhanden heisst nicht lesbar. Ein einzelnes kaputtes Symbol darf die
    Website nicht lahmlegen - hier als Ordner statt Datei nachgestellt, das
    verhaelt sich auf jedem Betriebssystem gleich."""
    ordner = tmp_path / "favicons"
    ordner.mkdir()
    (ordner / "a.ch.png").mkdir()
    monkeypatch.setattr(build, "FAVICON_DIR", ordner)
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    assert build.lade_favicons(build.load_data()) == {}


def test_icons_landen_im_gerenderten_html(mappe, monkeypatch, tmp_path):
    """Ende-zu-Ende: eine befuellte Icon-Tabelle muss auch tatsaechlich in
    der erzeugten Seite ankommen, nicht nur der Platzhalter im Rohtemplate."""
    ordner = tmp_path / "favicons"
    ordner.mkdir()
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk"
        "YPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
    )
    (ordner / "a.ch.png").write_bytes(png)
    monkeypatch.setattr(build, "FAVICON_DIR", ordner)
    pfad = mappe(LINK_HEADER, [SKILLS_ROW + ["https://a.ch/x", "", ""]])
    monkeypatch.setattr(build, "XLSX", pfad)
    ziel = tmp_path / "skillsliste.html"
    monkeypatch.setattr(build, "OUTPUT", ziel)

    data = build.load_data()
    icons = build.lade_favicons(data)
    build.render(data, icons)

    html = ziel.read_bytes().decode("utf-8-sig")
    assert '"a.ch": "data:image/png;base64,' in html
    assert re.search(r'var ICONS = \{"a\.ch": "data:image/png;base64,[^"]+"\};', html)
