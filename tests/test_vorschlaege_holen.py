import json
import os
import stat
import subprocess
import sys
import types
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "tools"))

import vorschlaege_holen as vh

BEISPIEL = {
    "art": "neu",
    "stufe": "Hoch",
    "kategorie": "Ablenkung",
    "emoji": "🎧",
    "titel": "Musik hören",
    "beschreibung": "Ein Lied auflegen.",
    "tipp": "Kopfhörer bereitlegen",
    "von": "Max",
}


def issue_text(nutzlast):
    return (
        "| Feld | Wert |\n| --- | --- |\n| Titel | Musik hören |\n\n"
        "<!-- vorschlag\n" + json.dumps(nutzlast, ensure_ascii=False) + "\n-->\n"
    )


def test_parse_body_liest_den_block():
    assert vh.parse_body(issue_text(BEISPIEL)) == BEISPIEL


def test_parse_body_ohne_block_gibt_none():
    assert vh.parse_body("Nur Text, kein Block.") is None


def test_parse_body_bei_kaputtem_json_gibt_none():
    assert vh.parse_body("<!-- vorschlag\n{kaputt\n-->") is None


def test_parse_body_lehnt_zwei_bloecke_ab():
    # issue_text() nimmt die Beschreibung nicht in die Tabelle auf, darum wird
    # der Rumpf hier von Hand gebaut – mit dem gefaelschten Block zuerst (so wie
    # ihn jemand in ein Freitextfeld schreiben wuerde) und dem echten Block
    # danach, genau wie es der Worker mit Tabelle + echtem Block tut.
    gift = '<!-- vorschlag {"art":"neu","titel":"EINGESCHLEUST"} -->'
    body = (
        "| Feld | Wert |\n| --- | --- |\n"
        f"| Beschreibung | Harmlos. {gift} |\n\n"
        "<!-- vorschlag\n" + json.dumps(BEISPIEL, ensure_ascii=False) + "\n-->\n"
    )
    assert vh.parse_body(body) is None


def test_anhaengen_schreibt_in_die_richtigen_spalten(tmp_path):
    # Kopfzeile bewusst in anderer Reihenfolge als die interne SPALTEN-Liste,
    # damit eine Rueckkehr zu positionsbasiertem Schreiben auffliegt.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["Titel", "Stufe", "Von", "Kategorie", "Beschreibung", "Emoji", "Tipp"])
    ws.append(["Vorhanden", "Hoch", "", "Ablenkung", "Alte Zeile", "🌶️", ""])
    wb.save(pfad)

    anzahl = vh.in_excel_uebernehmen(pfad, [], [BEISPIEL])

    assert anzahl == 1
    ws2 = openpyxl.load_workbook(pfad)["Skills"]
    assert ws2.max_row == 3
    kopf = [c.value for c in ws2[1]]
    zeile = dict(zip(kopf, [c.value for c in ws2[3]]))
    assert zeile == {
        "Titel": "Musik hören",
        "Stufe": "Hoch",
        "Von": "Max",
        "Kategorie": "Ablenkung",
        "Beschreibung": "Ein Lied auflegen.",
        "Emoji": "🎧",
        "Tipp": "Kopfhörer bereitlegen",
    }


BESTAND = {
    "hoch": {"kategorien": [{"id": "ablenkung", "label": "Ablenkung", "skills": []}]},
    "mittel": {"kategorien": []},
    "tief": {"kategorien": []},
}


def test_pruefung_laesst_einen_gueltigen_eintrag_durch():
    assert vh.pruefe_eintrag(BEISPIEL, BESTAND) is None


def test_pruefung_erlaubt_fehlenden_tipp_und_namen():
    eintrag = {s: w for s, w in BEISPIEL.items() if s not in ("tipp", "von")}
    assert vh.pruefe_eintrag(eintrag, BESTAND) is None


def test_pruefung_lehnt_zu_langen_titel_ab():
    grund = vh.pruefe_eintrag({**BEISPIEL, "titel": "x" * 61}, BESTAND)
    assert grund and "Titel" in grund


def test_pruefung_erlaubt_genau_60_zeichen_im_titel():
    assert vh.pruefe_eintrag({**BEISPIEL, "titel": "x" * 60}, BESTAND) is None


def test_pruefung_lehnt_spitze_klammern_ab():
    gift = "harmlos</script><script>alert(1)</script>"
    grund = vh.pruefe_eintrag({**BEISPIEL, "beschreibung": gift}, BESTAND)
    assert grund == "Spitze Klammern sind nicht erlaubt."


def test_pruefung_lehnt_kommentarzeichen_ab():
    grund = vh.pruefe_eintrag({**BEISPIEL, "tipp": "a <!-- b"}, BESTAND)
    assert grund == "Kommentarzeichen sind nicht erlaubt."


def test_pruefung_lehnt_links_ab():
    grund = vh.pruefe_eintrag({**BEISPIEL, "beschreibung": "siehe HTTP://spam.example"}, BESTAND)
    assert grund == "Links sind nicht erlaubt."


def test_pruefung_lehnt_unbekannte_kategorie_ab():
    """Der Fall, der beim Umbenennen einer Kategorie real auftritt."""
    grund = vh.pruefe_eintrag({**BEISPIEL, "kategorie": "Erfunden"}, BESTAND)
    assert grund and "Kategorie" in grund and "Erfunden" in grund


def test_pruefung_lehnt_unbekannte_stufe_ab():
    grund = vh.pruefe_eintrag({**BEISPIEL, "stufe": "Sehr hoch"}, BESTAND)
    assert grund and "Stufe" in grund


def test_pruefung_nimmt_ein_zusammengesetztes_emoji_an():
    # 🧘‍♀️ ist EIN sichtbares Zeichen (Person + ZWJ + Geschlechtszeichen),
    # aber vier Codepunkte – genau der Fall, den die alte Grenze von 2
    # faelschlich blockiert hat.
    assert vh.pruefe_eintrag({**BEISPIEL, "emoji": "🧘‍♀️"}, BESTAND) is None


def test_pruefung_lehnt_eine_zu_lange_emoji_kette_ab():
    grund = vh.pruefe_eintrag({**BEISPIEL, "emoji": "🎧" * 20}, BESTAND)
    assert grund == "Emoji ist zu lang."


def test_pruefung_lehnt_fehlenden_pflichtschluessel_ab():
    ohne = {s: w for s, w in BEISPIEL.items() if s != "emoji"}
    grund = vh.pruefe_eintrag(ohne, BESTAND)
    assert grund and "Emoji" in grund


def test_pruefung_lehnt_leeren_pflichtwert_ab():
    grund = vh.pruefe_eintrag({**BEISPIEL, "titel": "   "}, BESTAND)
    assert grund and "Titel" in grund


def test_hat_label_erkennt_vorhandenes_label():
    issue = {"labels": [{"name": "in Pruefung"}, {"name": "freigegeben"}]}
    assert vh.hat_label(issue, "freigegeben") is True


def test_hat_label_erkennt_fehlendes_label():
    issue = {"labels": [{"name": "in Pruefung"}]}
    assert vh.hat_label(issue, "freigegeben") is False


def test_hat_label_bei_leerer_labelliste():
    issue = {"labels": []}
    assert vh.hat_label(issue, "freigegeben") is False


def test_hat_label_bei_fehlendem_labelschluessel():
    issue = {}
    assert vh.hat_label(issue, "freigegeben") is False


def test_anhaengen_meldet_verstaendlich_wenn_datei_gesperrt(tmp_path):
    # Simuliert eine in Excel geoeffnete Datei: schreibgeschuetzt gesetzt,
    # damit wb.save() ein PermissionError wirft, genau wie beim echten Sperren.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp", "Von"])
    wb.save(pfad)

    os.chmod(pfad, stat.S_IREAD)
    try:
        with pytest.raises(SystemExit) as ausnahme:
            vh.in_excel_uebernehmen(pfad, [], [BEISPIEL])
        meldung = str(ausnahme.value)
        assert "laesst sich nicht speichern" in meldung
        assert "Excel" in meldung
        assert "nichts veraendert" in meldung
    finally:
        os.chmod(pfad, stat.S_IWRITE | stat.S_IREAD)


def test_anhaengen_legt_fehlende_von_spalte_an(tmp_path):
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp"])
    wb.save(pfad)

    vh.in_excel_uebernehmen(pfad, [], [BEISPIEL])

    ws2 = openpyxl.load_workbook(pfad)["Skills"]
    assert [c.value for c in ws2[1]][6] == "Von"
    assert [c.value for c in ws2[2]][6] == "Max"


def test_bereinigt_entfernt_randweissraum_der_die_pruefung_besteht():
    # Genau der gemeldete Fall: 60 Zeichen plus zwei Leerzeichen bestehen die
    # Pruefung (die auf der getrimmten Fassung prueft), duerfen aber nicht mit
    # 62 Zeichen in der Excel landen.
    eintrag = {**BEISPIEL, "titel": "x" * 60 + "  "}
    assert vh.pruefe_eintrag(eintrag, BESTAND) is None
    assert vh.bereinigt(eintrag)["titel"] == "x" * 60


def test_bereinigt_trimmt_alle_excel_spalten():
    eintrag = {**BEISPIEL, "beschreibung": "  Ein Lied auflegen.  ", "von": " Max "}
    ergebnis = vh.bereinigt(eintrag)
    assert ergebnis["beschreibung"] == "Ein Lied auflegen."
    assert ergebnis["von"] == "Max"


def freigegebenes_issue(nummer: int, daten: dict = BEISPIEL) -> dict:
    """Baut ein Issue, wie main() es von hole_issues() bekaeme – freigegeben,
    vom Formular-Konto, mit lesbarem Vorschlagsblock."""
    return {
        "number": nummer,
        "title": daten.get("titel", "t"),
        "author": {"login": vh.BOT},
        "labels": [{"name": vh.LABEL}],
        "body": issue_text(daten),
    }


def test_main_laesst_unerwartete_fehler_beim_schliessen_durchschlagen(monkeypatch):
    # except Exception war zu breit: ein Programmierfehler beim Schliessen sah
    # damit aus wie ein gescheiterter Netzaufruf. Nur subprocess.CalledProcessError
    # (der tatsaechliche Fehlerfall von issue_schliessen mit check=True) darf
    # abgefangen werden – alles andere muss durchschlagen.
    monkeypatch.setattr(vh, "hole_issues", lambda: [freigegebenes_issue(1)])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    # main() schreibt ueber genau eine Funktion – sie muss hier stillgelegt
    # werden, sonst liefe der Test in die echte skills_daten.xlsx.
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 0)

    def schlaegt_unerwartet_fehl(nummer):
        raise RuntimeError("Programmierfehler")

    monkeypatch.setattr(vh, "issue_schliessen", schlaegt_unerwartet_fehl)

    with pytest.raises(RuntimeError):
        vh.main()


def test_main_faengt_calledprocesserror_ab_und_zeigt_keine_widerspruechliche_erfolgszeile(
    monkeypatch, capsys
):
    monkeypatch.setattr(vh, "hole_issues", lambda: [freigegebenes_issue(1)])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    # 0 zurueckgeben, um in diesem Test den Skillsliste-Neubau (subprocess) nicht
    # auszuloesen – der Fokus liegt hier auf dem Schliessen-Fehler und der
    # Erfolgszeile, nicht auf build.py.
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 0)

    def schlaegt_wie_gh_fehl(nummer):
        raise subprocess.CalledProcessError(1, ["gh"])

    monkeypatch.setattr(vh, "issue_schliessen", schlaegt_wie_gh_fehl)

    vh.main()

    ausgabe = capsys.readouterr().out
    assert "#1" in ausgabe
    assert "konnte(n) nicht geschlossen werden" in ausgabe
    # Bei 0 tatsaechlich uebernommenen Zeilen darf kein gruenes Haekchen mit
    # einer Null erscheinen – widerspruechlich neben der Warnung direkt darunter.
    assert "✅" not in ausgabe


def test_main_gibt_keine_erfolgszeile_bei_null_uebernahmen_aus(monkeypatch, capsys):
    abgelehntes_issue = freigegebenes_issue(1, {**BEISPIEL, "kategorie": "Erfunden"})
    monkeypatch.setattr(vh, "hole_issues", lambda: [abgelehntes_issue])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 0)
    # Ohne diese Attrappe wuerde main() fuer den automatisch abgelehnten
    # Vorschlag echtes input() aufrufen – hier nicht das Testziel.
    monkeypatch.setattr(vh, "automatische_ablehnungen_melden", lambda faelle: [])

    vh.main()

    ausgabe = capsys.readouterr().out
    assert "✅" not in ausgabe
    assert "nicht übernommen" in ausgabe


def test_main_verwendet_einzahl_und_mehrzahl_korrekt(monkeypatch, capsys):
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    # main() ruft nach einer Uebernahme build.py auf; hier nur die Ausgabe der
    # Erfolgszeile pruefen, ohne einen echten build.py-Lauf auszuloesen.
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )

    monkeypatch.setattr(vh, "hole_issues", lambda: [freigegebenes_issue(1)])
    monkeypatch.setattr(
        vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: len(aenderungen) + len(neue)
    )
    vh.main()
    einzahl = capsys.readouterr().out
    assert "1 Vorschlag in" in einzahl
    assert "1 Vorschläge" not in einzahl

    monkeypatch.setattr(
        vh, "hole_issues", lambda: [freigegebenes_issue(1), freigegebenes_issue(2)]
    )
    vh.main()
    mehrzahl = capsys.readouterr().out
    assert "2 Vorschläge in" in mehrzahl


# ── Ausbaustufe 2: Änderungen an bestehenden Skills ──────────────────────────

AENDERUNG = {
    "art": "aenderung",
    "stufe": "Hoch",
    "kategorie": "Ablenkung",
    "original": "Musik hören",
    "emoji": "🎵",
    "titel": "Musik bewusst hören",
    "beschreibung": "Ein Lied aussuchen und nur darauf achten.",
    "tipp": "Kopfhörer bereitlegen",
    "erg": "Lea",
}

DATEN_MIT_SKILL = {
    "hoch": {"kategorien": [{"id": "ablenkung", "label": "Ablenkung", "skills": [
        {"e": "🎧", "t": "Musik hören", "b": "Ein Lied auflegen.", "tip": "", "von": "Max", "erg": ""}
    ]}]},
    "mittel": {"kategorien": []},
    "tief": {"kategorien": []},
}

KOPF = ["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp", "Von", "Ergaenzt"]


def mappe_mit_skill(tmp_path):
    """Eine Excel mit genau dem Skill, den DATEN_MIT_SKILL beschreibt."""
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(KOPF)
    ws.append(["Hoch", "Ablenkung", "🎧", "Musik hören", "Ein Lied auflegen.", "", "Max", ""])
    ws.append(["Tief", "Ruhe", "🌊", "Atmen", "Ruhig atmen.", "", "", ""])
    wb.save(pfad)
    return pfad


def test_aenderung_wird_angenommen():
    assert vh.pruefe_eintrag(AENDERUNG, DATEN_MIT_SKILL) is None


def test_aenderung_ohne_original_wird_abgelehnt():
    ohne = {k: v for k, v in AENDERUNG.items() if k != "original"}
    meldung = vh.pruefe_eintrag(ohne, DATEN_MIT_SKILL)
    assert meldung is not None


def test_aenderung_an_verschwundenem_skill_wird_abgelehnt():
    meldung = vh.pruefe_eintrag(dict(AENDERUNG, original="Gibt es nicht"), DATEN_MIT_SKILL)
    assert meldung is not None
    assert "nicht" in meldung.lower()


def test_aenderung_ersetzt_die_zeile(tmp_path):
    pfad = mappe_mit_skill(tmp_path)
    anzahl = vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])
    assert anzahl == 1
    ws = openpyxl.load_workbook(pfad)["Skills"]
    assert ws.max_row == 3, "es darf keine Zeile dazugekommen sein"
    kopf = [c.value for c in ws[1]]
    zeile = dict(zip(kopf, [c.value for c in ws[2]]))
    assert zeile["Titel"] == "Musik bewusst hören"
    assert zeile["Emoji"] == "🎵"
    assert zeile["Beschreibung"] == "Ein Lied aussuchen und nur darauf achten."
    assert zeile["Tipp"] == "Kopfhörer bereitlegen"
    assert zeile["Von"] == "Max", "der urspruengliche Beitragende bleibt stehen"
    assert zeile["Ergaenzt"] == "Lea"


def test_aenderung_ohne_passende_zeile_meldet_es(tmp_path):
    pfad = mappe_mit_skill(tmp_path)
    try:
        vh.in_excel_uebernehmen(pfad, [dict(AENDERUNG, original="Gibt es nicht")], [])
    except vh.ZeileNichtGefunden as fehler:
        assert "Gibt es nicht" in str(fehler)
    else:
        raise AssertionError("ZeileNichtGefunden erwartet")


def test_aenderung_erhaelt_filter_und_dropdown(tmp_path):
    # Beim Ersetzen einer Zeile darf weder der Filterbereich noch das
    # Stufen-Dropdown verlorengehen – beides sieht man der Datei nicht an,
    # es faellt erst auf, wenn jemand in Excel sortiert.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(KOPF)
    ws.append(["Hoch", "Ablenkung", "🎧", "Musik hören", "Ein Lied auflegen.", "", "Max", ""])
    ws.append(["Tief", "Ruhe", "🌊", "Atmen", "Ruhig atmen.", "", "", ""])
    ws.auto_filter.ref = "A1:H3"
    pruefung = DataValidation(type="list", formula1='"Hoch,Mittel,Tief"', allow_blank=True)
    ws.add_data_validation(pruefung)
    pruefung.add("A2:A3")
    wb.save(pfad)

    vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])

    ws2 = openpyxl.load_workbook(pfad)["Skills"]
    assert ws2.auto_filter.ref == "A1:H3"
    bereiche = [str(p.sqref) for p in ws2.data_validations.dataValidation]
    assert bereiche == ["A2:A3"]


def test_aenderung_zieht_den_filter_ueber_eine_neu_angelegte_spalte(tmp_path):
    # Aeltere Mappen haben die Spalte `Ergaenzt` noch nicht. Sie wird angelegt –
    # dann muss der Filter sie auch abdecken, sonst faellt sie beim Sortieren
    # heraus und die Werte verrutschen gegenueber den uebrigen Spalten.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(KOPF[:-1])
    ws.append(["Hoch", "Ablenkung", "🎧", "Musik hören", "Ein Lied auflegen.", "", "Max"])
    ws.append(["Tief", "Ruhe", "🌊", "Atmen", "Ruhig atmen.", "", ""])
    ws.auto_filter.ref = "A1:G3"
    pruefung = DataValidation(type="list", formula1='"Hoch,Mittel,Tief"', allow_blank=True)
    ws.add_data_validation(pruefung)
    pruefung.add("A2:A3")
    wb.save(pfad)

    vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])

    ws2 = openpyxl.load_workbook(pfad)["Skills"]
    assert [c.value for c in ws2[1]][7] == "Ergaenzt"
    assert [c.value for c in ws2[2]][7] == "Lea"
    assert ws2.auto_filter.ref == "A1:H3"
    assert [str(p.sqref) for p in ws2.data_validations.dataValidation] == ["A2:A3"]


def test_aenderung_bei_doppeltem_titel_aendert_nichts(tmp_path):
    # Stufe + Kategorie + Titel sind der Schluessel. Kommt er zweimal vor, waere
    # jede Wahl geraten – dann lieber gar nichts schreiben und nachfragen.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(KOPF)
    ws.append(["Hoch", "Ablenkung", "🎧", "Musik hören", "Ein Lied auflegen.", "", "Max", ""])
    ws.append(["Hoch", "Ablenkung", "🎼", "Musik hören", "Etwas anderes.", "", "Ida", ""])
    wb.save(pfad)

    with pytest.raises(vh.ZeileMehrdeutig) as ausnahme:
        vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])
    assert "Musik hören" in str(ausnahme.value)

    ws2 = openpyxl.load_workbook(pfad)["Skills"]
    assert [c.value for c in ws2[2]][2] == "🎧", "keine Zeile darf angefasst worden sein"
    assert [c.value for c in ws2[3]][2] == "🎼"


def test_aenderung_meldet_fehlende_schluesselspalte(tmp_path):
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["Stufe", "Kategorie", "Emoji", "Beschreibung", "Tipp", "Von"])
    ws.append(["Hoch", "Ablenkung", "🎧", "Ein Lied auflegen.", "", "Max"])
    wb.save(pfad)

    with pytest.raises(SystemExit) as ausnahme:
        vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])
    meldung = str(ausnahme.value)
    assert "Titel" in meldung
    assert "Kopfzeile" in meldung


def test_aenderung_meldet_verstaendlich_wenn_datei_gesperrt(tmp_path):
    # Gleiche Behandlung wie beim Anhaengen: eine in Excel geoeffnete Datei darf
    # keinen Absturz erzeugen, sondern muss sagen, was zu tun ist.
    pfad = mappe_mit_skill(tmp_path)
    os.chmod(pfad, stat.S_IREAD)
    try:
        with pytest.raises(SystemExit) as ausnahme:
            vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])
        meldung = str(ausnahme.value)
        assert "laesst sich nicht speichern" in meldung
        assert "Excel" in meldung
        assert "nichts veraendert" in meldung
    finally:
        os.chmod(pfad, stat.S_IWRITE | stat.S_IREAD)


NEUER_SKILL = {**BEISPIEL, "titel": "Spazieren gehen", "emoji": "🚶"}


def test_main_uebergibt_beide_arten_in_einem_schreibvorgang(monkeypatch, capsys):
    # Ein einziger Schreibvorgang fuer den ganzen Lauf: er gelingt oder er
    # scheitert – es gibt kein „halb uebernommen".
    aufrufe = []
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, NEUER_SKILL),
        freigegebenes_issue(2, AENDERUNG),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )

    def uebernehmen(pfad, aenderungen, neue):
        aufrufe.append(
            ([e["titel"] for e in aenderungen], [e["titel"] for e in neue])
        )
        return len(aenderungen) + len(neue)

    monkeypatch.setattr(vh, "in_excel_uebernehmen", uebernehmen)

    vh.main()

    assert aufrufe == [(["Musik bewusst hören"], ["Spazieren gehen"])]
    ausgabe = capsys.readouterr().out
    assert "~ Hoch / Ablenkung: Musik bewusst hören" in ausgabe
    assert "+ Hoch / Ablenkung: Spazieren gehen" in ausgabe
    assert "2 Vorschläge" in ausgabe


def test_main_schliesst_kein_issue_wenn_eine_aenderung_nicht_zuzuordnen_ist(monkeypatch):
    # Geschriebene Excel bei offenen Issues waere der Dublettenfall. Bricht der
    # Schreibvorgang ab, darf deshalb auch kein Issue geschlossen werden.
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, NEUER_SKILL),
        freigegebenes_issue(2, AENDERUNG),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    geschlossen = []
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: geschlossen.append(nummer))

    def uebernehmen(pfad, aenderungen, neue):
        raise vh.ZeileNichtGefunden("'Musik hören' in Hoch / Ablenkung")

    monkeypatch.setattr(vh, "in_excel_uebernehmen", uebernehmen)

    with pytest.raises(SystemExit) as ausnahme:
        vh.main()

    meldung = str(ausnahme.value)
    assert "Musik hören" in meldung
    assert "build.bat" in meldung, "die haeufigste Ursache muss genannt sein"
    assert "nichts veraendert" in meldung
    assert "freigegeben" in meldung
    assert geschlossen == [], "kein Issue darf geschlossen worden sein"


ZWEITE_AENDERUNG = {
    "art": "aenderung",
    "stufe": "Tief",
    "kategorie": "Ruhe",
    "original": "Atmen",
    "emoji": "💨",
    "titel": "Langsam atmen",
    "beschreibung": "Viermal ein, viermal aus.",
    "tipp": "",
    "erg": "Ida",
}


def inhalt(pfad):
    """Alle Zellen des Blattes `Skills` – zum Vergleich vorher/nachher."""
    ws = openpyxl.load_workbook(pfad)["Skills"]
    return [[c.value for c in zeile] for zeile in ws.iter_rows()]


def test_uebernahme_schreibt_nichts_wenn_eine_aenderung_nicht_passt(tmp_path):
    # Die wichtigste Eigenschaft: gespeichert wird erst, wenn ALLES zugeordnet
    # ist. Zwei zuordenbare Aenderungen, eine nicht zuordenbare und ein neuer
    # Skill – danach darf in der Datei nichts davon stehen, auch nicht die
    # beiden Aenderungen, die fuer sich genommen gepasst haetten.
    pfad = mappe_mit_skill(tmp_path)
    vorher = inhalt(pfad)

    with pytest.raises(vh.ZeileNichtGefunden):
        vh.in_excel_uebernehmen(
            pfad,
            [AENDERUNG, ZWEITE_AENDERUNG, dict(AENDERUNG, original="Gibt es nicht")],
            [NEUER_SKILL],
        )

    assert inhalt(pfad) == vorher, "die Datei muss unveraendert geblieben sein"


def test_uebernahme_speichert_genau_einmal(tmp_path, monkeypatch):
    # Die Zusage „Es wurde nichts veraendert" haelt nur, solange es EINEN
    # Speichervorgang gibt. Ein zweiter koennte scheitern, nachdem der erste
    # geschrieben hat – dann waere die Excel halb uebernommen und die Meldung
    # gelogen.
    #
    # Gezaehlt wird das ECHTE Schreiben (openpyxl) und nicht die Hilfsfunktion
    # speichern(): ein zusaetzliches wb.save() daneben wuerde diesem Test sonst
    # entgehen.
    pfad = mappe_mit_skill(tmp_path)
    echtes_speichern = openpyxl.Workbook.save
    aufrufe = []

    def gezaehlt(self, ziel):
        aufrufe.append(str(ziel))
        return echtes_speichern(self, ziel)

    monkeypatch.setattr(openpyxl.Workbook, "save", gezaehlt)

    vh.in_excel_uebernehmen(pfad, [AENDERUNG, ZWEITE_AENDERUNG], [NEUER_SKILL])

    assert len(aufrufe) == 1, "es darf genau einmal gespeichert werden"
    ws = openpyxl.load_workbook(pfad)["Skills"]
    assert ws.max_row == 4
    assert ws["D2"].value == "Musik bewusst hören"
    assert ws["D3"].value == "Langsam atmen"
    assert ws["D4"].value == "Spazieren gehen"


def test_uebernahme_schreibt_nichts_wenn_die_datei_gesperrt_ist(tmp_path):
    # Ein einziger Speichervorgang fuer Aenderungen UND neue Zeilen: scheitert
    # er, ist die Zusage „Es wurde nichts veraendert" in jedem Fall wahr.
    pfad = mappe_mit_skill(tmp_path)
    vorher = inhalt(pfad)
    os.chmod(pfad, stat.S_IREAD)
    try:
        with pytest.raises(SystemExit) as ausnahme:
            vh.in_excel_uebernehmen(pfad, [AENDERUNG], [NEUER_SKILL])
        assert "nichts veraendert" in str(ausnahme.value)
    finally:
        os.chmod(pfad, stat.S_IWRITE | stat.S_IREAD)

    assert inhalt(pfad) == vorher, "die Datei muss unveraendert geblieben sein"


def test_abbruch_mitten_im_schreiben_laesst_die_datei_heil(tmp_path, monkeypatch):
    # Der schlimmste Fall: volle Festplatte oder weggebrochenes Laufwerk mitten
    # im Schreiben. Die Excel ist das Original, aus dem alles andere erzeugt
    # wird – ein Truemmerstand waere nicht wiederherstellbar. Darum wird in eine
    # Nachbardatei geschrieben und erst am Ende umgelegt.
    pfad = mappe_mit_skill(tmp_path)
    vorher = pfad.read_bytes()

    def bricht_mittendrin_ab(self, ziel):
        Path(ziel).write_bytes(b"halb geschriebene Truemmer")
        raise OSError(28, "No space left on device")

    monkeypatch.setattr(openpyxl.Workbook, "save", bricht_mittendrin_ab)

    with pytest.raises(SystemExit) as ausnahme:
        vh.in_excel_uebernehmen(pfad, [AENDERUNG], [NEUER_SKILL])

    meldung = str(ausnahme.value)
    assert "nicht geschrieben werden" in meldung
    assert "unveraendert" in meldung
    assert pfad.read_bytes() == vorher, "die Originaldatei muss unberuehrt sein"
    openpyxl.load_workbook(pfad)  # muss weiterhin lesbar sein
    assert [p.name for p in tmp_path.iterdir()] == [pfad.name], (
        "es darf keine halb geschriebene Nachbardatei liegen bleiben"
    )


def test_abbruch_beim_umlegen_laesst_die_datei_heil(tmp_path, monkeypatch):
    # Die Nachbardatei wird UMGELEGT, nicht kopiert: os.replace ist ein einziger
    # Schritt. Ein Kopieren saehe im Erfolgsfall gleich aus, koennte aber
    # mittendrin abbrechen und wieder eine halbe Zieldatei hinterlassen.
    pfad = mappe_mit_skill(tmp_path)
    vorher = pfad.read_bytes()

    def scheitert(quelle, ziel):
        raise OSError(28, "No space left on device")

    monkeypatch.setattr(vh.os, "replace", scheitert)

    with pytest.raises(SystemExit) as ausnahme:
        vh.in_excel_uebernehmen(pfad, [AENDERUNG], [NEUER_SKILL])

    assert "nicht geschrieben werden" in str(ausnahme.value)
    assert pfad.read_bytes() == vorher
    assert [p.name for p in tmp_path.iterdir()] == [pfad.name]


def test_die_neue_datei_steht_vor_dem_umlegen_auf_der_platte(tmp_path, monkeypatch):
    # Zwischen „geschrieben" und „auf der Platte" liegt der Schreibpuffer des
    # Betriebssystems. Ohne fsync koennte ein Stromausfall kurz nach dem Umlegen
    # eine abgeschnittene Mappe hinterlassen – die Datei ist nicht ersetzbar.
    pfad = mappe_mit_skill(tmp_path)
    schritte = []
    echtes_fsync, echtes_replace = vh.os.fsync, vh.os.replace

    def gemerkt_fsync(nummer):
        schritte.append("fsync")
        return echtes_fsync(nummer)

    def gemerkt_replace(quelle, ziel):
        schritte.append("replace")
        return echtes_replace(quelle, ziel)

    monkeypatch.setattr(vh.os, "fsync", gemerkt_fsync)
    monkeypatch.setattr(vh.os, "replace", gemerkt_replace)

    vh.in_excel_uebernehmen(pfad, [AENDERUNG], [])

    assert schritte == ["fsync", "replace"]


def test_steuerzeichen_werden_still_entfernt(tmp_path):
    # Aus Word oder einem PDF kopierter Text enthaelt manchmal unsichtbare
    # Steuerzeichen. Excel kann sie nicht speichern. Sie tragen keine Bedeutung,
    # darum werden sie entfernt statt den Vorschlag abzulehnen – mit einer
    # Meldung „Steuerzeichen gefunden" koennte niemand etwas anfangen.
    eintrag = {**NEUER_SKILL, "beschreibung": "Aus Word\x07 kopiert.\x0b"}
    assert vh.pruefe_eintrag(eintrag, DATEN_MIT_SKILL) is None

    pfad = mappe_mit_skill(tmp_path)
    vh.in_excel_uebernehmen(pfad, [], [vh.bereinigt(eintrag)])

    ws = openpyxl.load_workbook(pfad)["Skills"]
    kopf = [c.value for c in ws[1]]
    zeile = dict(zip(kopf, [c.value for c in ws[ws.max_row]]))
    assert zeile["Beschreibung"] == "Aus Word kopiert."


def test_zeilenumbruch_und_tabulator_bleiben_erhalten():
    # Umbruch und Tabulator sind sichtbarer Text, keine Stoerzeichen – Excel
    # speichert beide anstandslos. Sie duerfen beim Saeubern nicht mitgehen.
    eintrag = {**NEUER_SKILL, "beschreibung": "eins\nzwei\tdrei"}
    assert vh.bereinigt(eintrag)["beschreibung"] == "eins\nzwei\tdrei"


def test_anhaengen_zieht_filter_und_dropdown_auf_die_neue_zeile(tmp_path):
    # Fallen die neuen Zeilen aus Filterbereich und Dropdown heraus, verrutschen
    # beim Sortieren in Excel die Werte gegeneinander – stille Verfaelschung.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(KOPF)
    ws.append(["Hoch", "Ablenkung", "🎧", "Musik hören", "Ein Lied auflegen.", "", "Max", ""])
    ws.append(["Tief", "Ruhe", "🌊", "Atmen", "Ruhig atmen.", "", "", ""])
    ws.auto_filter.ref = "A1:H3"
    pruefung = DataValidation(type="list", formula1='"Hoch,Mittel,Tief"', allow_blank=True)
    ws.add_data_validation(pruefung)
    pruefung.add("A2:A3")
    wb.save(pfad)

    vh.in_excel_uebernehmen(pfad, [], [NEUER_SKILL])

    ws2 = openpyxl.load_workbook(pfad)["Skills"]
    assert ws2.max_row == 4
    assert ws2.auto_filter.ref == "A1:H4"
    assert [str(p.sqref) for p in ws2.data_validations.dataValidation] == ["A2:A4"]


def test_uebernahme_meldet_fehlende_mappe(tmp_path):
    with pytest.raises(SystemExit) as ausnahme:
        vh.in_excel_uebernehmen(tmp_path / "gibt_es_nicht.xlsx", [], [NEUER_SKILL])
    meldung = str(ausnahme.value)
    assert "gibt_es_nicht.xlsx" in meldung
    assert "nicht gefunden" in meldung


def test_uebernahme_meldet_umbenanntes_blatt(tmp_path):
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills alt"
    ws.append(KOPF)
    wb.save(pfad)

    with pytest.raises(SystemExit) as ausnahme:
        vh.in_excel_uebernehmen(pfad, [], [NEUER_SKILL])
    meldung = str(ausnahme.value)
    assert "Skills" in meldung
    assert "umbenannt" in meldung or "nicht umbenennen" in meldung


def test_uebernahme_meldet_unlesbare_datei(tmp_path):
    pfad = tmp_path / "skills_daten.xlsx"
    pfad.write_bytes(b"das ist keine Excel-Datei")

    with pytest.raises(SystemExit) as ausnahme:
        vh.in_excel_uebernehmen(pfad, [], [NEUER_SKILL])
    assert "nicht lesen" in str(ausnahme.value)


def test_main_warnt_vor_offenen_issues_bevor_ein_fehler_durchschlaegt(monkeypatch, capsys):
    # Nach dem Schreiben ist die Excel voll, aber die Issues sind noch offen.
    # Schlaegt hier ein unerwarteter Fehler durch, muss die Warnung trotzdem
    # erscheinen – sonst weiss niemand, dass beim naechsten Lauf Dubletten
    # drohen.
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, NEUER_SKILL),
        freigegebenes_issue(2, {**NEUER_SKILL, "titel": "Zweiter"}),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 2)

    def schlaegt_unerwartet_fehl(nummer):
        raise RuntimeError("Programmierfehler")

    monkeypatch.setattr(vh, "issue_schliessen", schlaegt_unerwartet_fehl)

    with pytest.raises(RuntimeError):
        vh.main()

    ausgabe = capsys.readouterr().out
    assert "#1" in ausgabe and "#2" in ausgabe
    assert "bereits in der Excel" in ausgabe
    assert "von Hand" in ausgabe


def test_main_sagt_was_mit_offen_gebliebenen_issues_zu_tun_ist(monkeypatch, capsys):
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, {**BEISPIEL, "kategorie": "Erfunden"}),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 0)
    # Ohne diese Attrappe wuerde main() fuer den automatisch abgelehnten
    # Vorschlag echtes input() aufrufen – hier nicht das Testziel.
    monkeypatch.setattr(vh, "automatische_ablehnungen_melden", lambda faelle: [])

    vh.main()

    ausgabe = capsys.readouterr().out
    assert "nicht übernommen" in ausgabe
    # Eine Meldung ohne Handlungsanweisung laesst den Menschen ratlos zurueck.
    assert "abgelehnt" in ausgabe
    assert "github.com" in ausgabe


def test_aenderung_ohne_urspruenglichen_titel_trifft_keine_leere_zeile(tmp_path):
    # Ein leerer urspruenglicher Titel wuerde sonst auf eine Zeile mit leerer
    # Titelzelle passen und die falsche Zeile ueberschreiben. pruefe_eintrag
    # faengt das ab – die Sperre gehoert trotzdem direkt vor das Schreiben.
    pfad = tmp_path / "skills_daten.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(KOPF)
    ws.append(["Hoch", "Ablenkung", "🎧", "", "Zeile ohne Titel.", "", "Max", ""])
    wb.save(pfad)
    vorher = inhalt(pfad)

    with pytest.raises(vh.ZeileNichtGefunden):
        vh.in_excel_uebernehmen(pfad, [dict(AENDERUNG, original="")], [])

    assert inhalt(pfad) == vorher


def test_aenderung_greift_nicht_auf_eine_im_selben_lauf_neue_zeile(tmp_path):
    # Aenderungen werden VOR den neuen Zeilen eingearbeitet: eine Aenderung darf
    # sich nie auf einen Skill beziehen, den derselbe Lauf erst anlegt.
    pfad = mappe_mit_skill(tmp_path)
    vorher = inhalt(pfad)

    with pytest.raises(vh.ZeileNichtGefunden):
        vh.in_excel_uebernehmen(
            pfad, [dict(AENDERUNG, original="Spazieren gehen")], [NEUER_SKILL]
        )

    assert inhalt(pfad) == vorher


# ── Ausbaustufe 3: Duplikatpruefung mit Rueckfrage ───────────────────────────

TREFFER = [{
    "titel": "Lieblingslied auflegen", "aehnlich_zu": "Musik hören",
    "stufe": "Hoch", "kategorie": "Ablenkung",
    "begruendung": "Beide beschreiben gezieltes Musikhoeren.",
    "sicherheit": "unsicher",
}]


def uebernehmen_liste():
    return [({"number": 7}, {"art": "neu", "stufe": "Hoch", "kategorie": "Ablenkung",
                             "titel": "Lieblingslied auflegen", "emoji": "🎵",
                             "beschreibung": "Ein Lied aussuchen und nur darauf achten.",
                             "tipp": ""})]


def test_nachfrage_uebernehmen_laesst_den_vorschlag_drin(capsys):
    offen, ablehnungen = vh.nachfragen(TREFFER, uebernehmen_liste(), {}, eingabe=lambda _: "ü")
    assert offen == []
    assert ablehnungen == []
    ausgabe = capsys.readouterr().out
    assert "Musik hören" in ausgabe
    assert "Beide beschreiben" in ausgabe, "die Begruendung muss sichtbar sein"


def test_nachfrage_weiter_nimmt_den_vorschlag_heraus():
    raus, _ = vh.nachfragen(TREFFER, uebernehmen_liste(), {}, eingabe=lambda _: "w")
    assert raus == [7]


def test_nachfrage_akzeptiert_grossschreibung_und_leerzeichen():
    raus, _ = vh.nachfragen(TREFFER, uebernehmen_liste(), {}, eingabe=lambda _: " W ")
    assert raus == [7]


def test_nachfrage_fragt_erneut_bei_unsinniger_eingabe():
    antworten = iter(["x", "", "w"])
    raus, _ = vh.nachfragen(TREFFER, uebernehmen_liste(), {},
                            eingabe=lambda _: next(antworten))
    assert raus == [7]


def test_nachfrage_ohne_tastatur_ueberspringt_sicherheitshalber():
    """Laeuft das Skript ohne Eingabemoeglichkeit, wird NICHT stillschweigend
    uebernommen – lieber bleibt das Issue offen."""
    def keine_tastatur(_):
        raise EOFError

    raus, ablehnungen = vh.nachfragen(TREFFER, uebernehmen_liste(), {}, eingabe=keine_tastatur)
    assert raus == [7]
    assert ablehnungen == []


def test_treffer_ohne_passendes_issue_wird_ignoriert():
    """Nennt die KI einen Titel, den es hier gar nicht gibt, darf nichts passieren."""
    fremd = [dict(TREFFER[0], titel="Gibt es nicht")]
    raus, _ = vh.nachfragen(fremd, uebernehmen_liste(), {}, eingabe=lambda _: "w")
    assert raus == []


# ── Abschlusspruefung I-1: Rueckzuordnung eines Treffers zum richtigen Issue ──


def _neu(nummer, titel, stufe="Hoch", kategorie="Ablenkung"):
    return ({"number": nummer},
            {"art": "neu", "stufe": stufe, "kategorie": kategorie, "titel": titel,
             "emoji": "🙂", "beschreibung": f"Beschreibung von {titel}.", "tipp": ""})


def _aenderung(nummer, titel, stufe="Hoch", kategorie="Ablenkung"):
    return ({"number": nummer},
            {"art": "aenderung", "stufe": stufe, "kategorie": kategorie,
             "titel": titel, "original": titel})


def _treffer(titel, stufe="Hoch", kategorie="Ablenkung"):
    return [{"titel": titel, "aehnlich_zu": "Musik hören", "stufe": stufe,
             "kategorie": kategorie, "begruendung": "Dieselbe Handlung.",
             "sicherheit": "unsicher"}]


def test_nachfrage_trifft_nie_eine_aenderung():
    """Eine Aenderung behaelt im Regelfall ihren Titel – ein neuer Skill kann
    denselben tragen. Zugeordnet wurde frueher allein ueber den Titel, ueber
    ganz `uebernehmen`; im dict gewann der letzte Eintrag. Ergebnis: die
    Dublette wurde eingetragen und stattdessen die Aenderung herausgeworfen,
    die laut CLAUDE.md gar nicht geprueft wird. Geprueft werden nur neue
    Skills – eine Aenderung darf durch eine Rueckfrage NIEMALS herausfallen.
    """
    uebernehmen = [_neu(101, "Ammoniak riechen"), _aenderung(102, "Ammoniak riechen")]
    raus, _ = vh.nachfragen(_treffer("Ammoniak riechen"), uebernehmen, {},
                            eingabe=lambda _: "w")
    assert raus == [101], "der neue Skill muss raus, die Aenderung bleibt"


def test_nachfrage_zu_einer_aenderung_allein_fragt_gar_nicht():
    """Ohne einen passenden NEUEN Vorschlag gibt es nichts zu fragen."""
    uebernehmen = [_aenderung(102, "Ammoniak riechen")]
    raus, _ = vh.nachfragen(_treffer("Ammoniak riechen"), uebernehmen, {},
                            eingabe=lambda _: "w")
    assert raus == []


def test_nachfrage_unterscheidet_zwei_neue_mit_gleichem_titel():
    """Zwei neue Vorschlaege mit demselben Titel in verschiedenen Stufen:
    Stufe und Kategorie stehen im Treffer und machen die Zuordnung eindeutig."""
    uebernehmen = [
        _neu(1, "Kaltes Wasser", "Hoch", "Ablenkung"),
        _neu(2, "Kaltes Wasser", "Tief", "Ruhe"),
    ]
    raus, _ = vh.nachfragen(_treffer("Kaltes Wasser", "Tief", "Ruhe"), uebernehmen, {},
                            eingabe=lambda _: "w")
    assert raus == [2]


def test_nachfrage_findet_den_vorschlag_auch_bei_abweichender_stufe_im_treffer():
    """Rueckfallweg: Das Antwortschema kennt nur „Zeichenkette" – ob das Modell
    in `stufe`/`kategorie` den Vorschlag oder den aehnlichen Bestandsskill
    meint, laesst sich nicht erzwingen (der Prompt sagt es, mehr geht nicht).
    Nennt es die Werte des Bestandsskills, faende ein reiner
    Dreier-Schluessel nichts und die Rueckfrage entfiele lautlos. Solange der
    Titel unter den neuen Vorschlaegen eindeutig ist, greift deshalb er.
    """
    uebernehmen = [_neu(5, "Kaltes Wasser", "Hoch", "Ablenkung")]
    raus, _ = vh.nachfragen(_treffer("Kaltes Wasser", "Tief", "Ruhe"), uebernehmen, {},
                            eingabe=lambda _: "w")
    assert raus == [5]


def test_nachfrage_ueberspringt_einen_unbrauchbaren_treffer_und_fragt_zum_naechsten():
    """Zweite Schutzebene hinter dem Typfilter in `duplikat.pruefe_duplikate`:
    Ein Treffer mit falsch getyptem Titel waere als dict-Schluessel ein
    `TypeError: unhashable type`. Er faellt weg, ohne die Rueckfrage zu den
    uebrigen Treffern mitzureissen."""
    uebernehmen = [_neu(1, "Kaltes Wasser"), _neu(2, "Warme Dusche")]
    treffer = _treffer("Warme Dusche")
    treffer.insert(0, dict(treffer[0], titel=["Kaltes Wasser"]))
    raus, _ = vh.nachfragen(treffer, uebernehmen, {}, eingabe=lambda _: "w")
    assert raus == [2]


# ── Ausbaustufe 3 Nachbesserung: Verdrahtung in main() ───────────────────────
#
# nachfragen() allein zu testen deckt NICHT ab, dass main() die Pruefung
# ueberhaupt richtig aufruft: nur fuer neue Skills, nur mit Schluessel, und
# ohne dass ein werfendes client_bauen() den ganzen Lauf mitreisst. Genau das
# war die Luecke, die eine Pruefung an dieser Stelle fand (die Duplikatpruefung
# lief unter einer Mutation auch fuer Aenderungen mit, ohne dass ein Test rot
# wurde).

def test_main_ruft_duplikatpruefung_nur_fuer_neue_skills_auf(monkeypatch):
    """Eine Aenderung zeigt schon ueber Stufe/Kategorie/Titel auf einen
    bestimmten vorhandenen Skill – "aehnelt einem vorhandenen Skill" ist dort
    keine sinnvolle Frage. Nur neue Skills duerfen an pruefe_duplikate() gehen,
    auch wenn im selben Lauf beide Arten vorkommen."""
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, NEUER_SKILL),
        freigegebenes_issue(2, AENDERUNG),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )
    monkeypatch.setattr(
        vh, "in_excel_uebernehmen",
        lambda pfad, aenderungen, neue: len(aenderungen) + len(neue),
    )

    monkeypatch.setattr(vh.duplikat, "schluessel_finden", lambda projekt: "fake-schluessel")
    monkeypatch.setattr(vh.duplikat, "client_bauen", lambda schluessel: object())
    gesehen = []

    def gefaelscht(neue, bestand, client):
        gesehen.append([n["titel"] for n in neue])
        return []

    monkeypatch.setattr(vh.duplikat, "pruefe_duplikate", gefaelscht)

    vh.main()

    assert gesehen == [["Spazieren gehen"]], (
        "nur der neue Skill darf geprueft werden, die Aenderung nicht"
    )


def test_main_ruft_duplikatpruefung_nicht_ohne_neue_skills(monkeypatch):
    """Nur eine Aenderung im Lauf, kein neuer Skill: pruefe_duplikate() darf gar
    nicht erst aufgerufen werden – ohne neue Skills gibt es nichts zu pruefen.

    Absichtlich KEIN "wirft AssertionError, wenn aufgerufen": main() faengt seit
    dem Client-Bau-Fix (siehe test_main_faengt_einen_werfenden_client_bauen_ab)
    JEDE Ausnahme aus diesem Block ab und laeuft normal weiter – ein
    AssertionError als Wachposten wuerde also lautlos verschluckt und die
    Pruefung waere wertlos. Stattdessen wird der Aufruf aufgezeichnet und NACH
    main() geprueft.
    """
    monkeypatch.setattr(vh, "hole_issues", lambda: [freigegebenes_issue(2, AENDERUNG)])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )
    monkeypatch.setattr(
        vh, "in_excel_uebernehmen",
        lambda pfad, aenderungen, neue: len(aenderungen) + len(neue),
    )

    monkeypatch.setattr(vh.duplikat, "schluessel_finden", lambda projekt: "fake-schluessel")

    client_bauen_aufrufe = []
    monkeypatch.setattr(
        vh.duplikat, "client_bauen", lambda *a, **k: client_bauen_aufrufe.append(1)
    )
    pruefe_aufrufe = []
    monkeypatch.setattr(
        vh.duplikat, "pruefe_duplikate", lambda *a, **k: pruefe_aufrufe.append(1)
    )

    vh.main()

    assert client_bauen_aufrufe == [], "ohne neue Skills darf kein Client gebaut werden"
    assert pruefe_aufrufe == [], "ohne neue Skills darf nicht geprueft werden"


def test_main_ruft_duplikatpruefung_nicht_ohne_schluessel(monkeypatch):
    """Ohne Schluessel entfaellt die Pruefung kommentarlos – client_bauen() darf
    dann nicht einmal versucht werden. Ginge der Schluessel-Filter verloren,
    wuerde client_bauen() mit dem echten anthropic-Paket installiert einen
    echten Client bauen und pruefe_duplikate() einen echten Netzversuch
    ausloesen, obwohl kein Schluessel hinterlegt ist – hier unabhaengig davon
    festgenagelt, ob `anthropic` in der Testumgebung installiert ist oder
    nicht, weil client_bauen() direkt durch eine Attrappe ersetzt wird.

    Aufzeichnen statt werfen (siehe Begruendung in
    test_main_ruft_duplikatpruefung_nicht_ohne_neue_skills): main() faengt
    Ausnahmen aus diesem Block ab, ein Wachposten-AssertionError wuerde also
    nichts beweisen.
    """
    monkeypatch.setattr(vh, "hole_issues", lambda: [freigegebenes_issue(1, NEUER_SKILL)])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )
    monkeypatch.setattr(
        vh, "in_excel_uebernehmen",
        lambda pfad, aenderungen, neue: len(aenderungen) + len(neue),
    )

    monkeypatch.setattr(vh.duplikat, "schluessel_finden", lambda projekt: None)

    client_bauen_aufrufe = []
    monkeypatch.setattr(
        vh.duplikat, "client_bauen", lambda *a, **k: client_bauen_aufrufe.append(1)
    )

    vh.main()

    assert client_bauen_aufrufe == [], (
        "ohne Schluessel haette client_bauen() nicht gerufen werden duerfen"
    )


def test_main_faengt_einen_werfenden_client_bauen_ab(monkeypatch, capsys):
    """Wirft schon der Bau des Clients (ungueltiger Schluessel, fehlendes Paket,
    …), liegt das AUSSERHALB der eigenen Absicherung von pruefe_duplikate():
    client_bauen() wird ausgewertet, BEVOR pruefe_duplikate() ueberhaupt
    laeuft. Ohne ein eigenes Sicherheitsnetz in main() stuerzt der ganze Lauf
    ab – genau die Zusage, die Task 3 fuer pruefe_duplikate() schon abgesichert
    hat, waere eine Ebene hoeher wieder aufgerissen."""
    monkeypatch.setattr(vh, "hole_issues", lambda: [freigegebenes_issue(1, NEUER_SKILL)])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )
    aufrufe = []

    def uebernehmen(pfad, aenderungen, neue):
        aufrufe.append([n["titel"] for n in neue])
        return len(aenderungen) + len(neue)

    monkeypatch.setattr(vh, "in_excel_uebernehmen", uebernehmen)

    monkeypatch.setattr(vh.duplikat, "schluessel_finden", lambda projekt: "fake-schluessel")

    def platzt(schluessel):
        raise RuntimeError("kaputter Schluessel")

    monkeypatch.setattr(vh.duplikat, "client_bauen", platzt)

    # Wird client_bauen() aufgeloest, BEVOR pruefe_duplikate() aufgerufen wird
    # (die Reihenfolge, die der Fix verlangt), erreicht dieser Aufruf hier nie
    # etwas – weder in der reparierten noch in der urspruenglich fehlerhaften
    # Fassung, denn `client_bauen(schluessel)` wird als Ausdruck ausgewertet,
    # BEVOR der Funktionsaufruf von pruefe_duplikate() ueberhaupt beginnt.
    # Aufzeichnen statt werfen, aus demselben Grund wie in den beiden Tests
    # oben: main() faengt Ausnahmen aus diesem Block ab.
    pruefe_aufrufe = []
    monkeypatch.setattr(
        vh.duplikat, "pruefe_duplikate", lambda *a, **k: pruefe_aufrufe.append(1)
    )

    vh.main()  # darf NICHT platzen, obwohl client_bauen() wirft

    ausgabe = capsys.readouterr().out
    assert "uebersprungen" in ausgabe
    assert "RuntimeError" in ausgabe
    assert pruefe_aufrufe == [], "pruefe_duplikate() haette nie erreicht werden duerfen"
    assert aufrufe == [["Spazieren gehen"]], (
        "die Uebernahme muss trotzdem normal weiterlaufen und den Skill schreiben"
    )


# ── Abschlusspruefung C-1: auch das VERARBEITEN der Antwort ist abgesichert ──
#
# Die beiden bereits behobenen Stellen betrafen das BESCHAFFEN der Antwort
# (`antwort.content[0]`, `client_bauen`). Die dritte lag eine Ebene hoeher:
# alles, was NACH `pruefe_duplikate()` mit dem Ergebnis geschieht, stand
# ausserhalb jeder Absicherung.


def _duplikat_lauf(monkeypatch, issues, pruefe, aufrufe, eingabe=None):
    """Stellt main() so, dass nur die Duplikatpruefung interessant ist.

    `aufrufe` sammelt, was in_excel_uebernehmen zu sehen bekaeme – daran
    haengt die Zusicherung „die Uebernahme laeuft weiter".
    """
    monkeypatch.setattr(vh, "hole_issues", lambda: issues)
    monkeypatch.setattr(vh, "lade_datenstand", lambda: DATEN_MIT_SKILL)
    monkeypatch.setattr(vh, "issue_schliessen", lambda nummer: None)
    monkeypatch.setattr(
        vh.subprocess, "run", lambda *a, **k: types.SimpleNamespace(returncode=0)
    )

    def uebernehmen(pfad, aenderungen, neue):
        aufrufe.append(([d["titel"] for d in aenderungen], [d["titel"] for d in neue]))
        return len(aenderungen) + len(neue)

    monkeypatch.setattr(vh, "in_excel_uebernehmen", uebernehmen)
    monkeypatch.setattr(vh.duplikat, "schluessel_finden", lambda projekt: "fake-schluessel")
    monkeypatch.setattr(vh.duplikat, "client_bauen", lambda schluessel: object())
    monkeypatch.setattr(vh.duplikat, "pruefe_duplikate", pruefe)
    if eingabe is not None:
        echte = vh.nachfragen
        monkeypatch.setattr(
            vh, "nachfragen", lambda t, u, b: echte(t, u, b, eingabe=eingabe)
        )


def test_main_ueberlebt_treffer_mit_unbrauchbaren_feldern(monkeypatch, capsys):
    """Ein Treffer, dessen `titel` keine Zeichenkette ist, wurde als
    dict-Schluessel benutzt: `TypeError: unhashable type: 'list'`, roher
    Traceback, ganzer Uebernahmelauf tot. Hier bewusst an `pruefe_duplikate`
    vorbei eingespeist – der Filter dort ist die erste Abwehrlinie, diese
    Zusicherung muss aber auch ohne ihn halten."""
    aufrufe = []
    kaputt = [{
        "titel": ["Spazieren gehen"], "aehnlich_zu": "Musik hören",
        "stufe": "Hoch", "kategorie": "Ablenkung", "begruendung": "…",
    }]
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: kaputt, aufrufe,
    )

    vh.main()  # darf NICHT platzen

    assert aufrufe == [([], ["Spazieren gehen"])], (
        "die Uebernahme muss den Skill trotzdem schreiben"
    )


def test_main_ueberlebt_einen_fehler_beim_verarbeiten_der_treffer(monkeypatch, capsys):
    """Der try-Block umschloss nur das Beschaffen der Antwort. Alles, was
    danach mit ihr geschieht – die Rueckfrage und das Umsortieren – muss
    ebenfalls darin liegen, sonst reisst ein Fehler dort den ganzen Lauf mit.
    """
    aufrufe = []
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: TREFFER, aufrufe,
    )

    def platzt(treffer, uebernehmen, bestand):
        raise RuntimeError("beim Verarbeiten geplatzt")

    monkeypatch.setattr(vh, "nachfragen", platzt)

    vh.main()  # darf NICHT platzen

    ausgabe = capsys.readouterr().out
    assert "uebersprungen" in ausgabe
    assert "RuntimeError" in ausgabe
    assert aufrufe == [([], ["Spazieren gehen"])], (
        "die Uebernahme muss trotzdem normal weiterlaufen"
    )


# ── Abschlusspruefung I-1: die Rueckfrage trifft den richtigen Vorschlag ─────

AENDERUNG_GLEICHER_TITEL = {**AENDERUNG, "titel": "Musik hören"}


def test_main_wirft_bei_gleichem_titel_die_aenderung_nicht_heraus(monkeypatch, capsys):
    """Ein neuer Skill und eine Aenderung mit demselben Titel: Die KI meldet
    den NEUEN Skill, der Mensch waehlt „weiter". Frueher wurde der neue Skill
    trotzdem eingetragen und stattdessen die Aenderung verworfen – das
    Gegenteil der Anweisung, und die Dublette landete in der Mappe."""
    aufrufe = []
    _duplikat_lauf(
        monkeypatch,
        [freigegebenes_issue(1, BEISPIEL), freigegebenes_issue(2, AENDERUNG_GLEICHER_TITEL)],
        lambda neue, bestand, client: [{
            "titel": "Musik hören", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: "w",
    )

    vh.main()

    assert aufrufe == [(["Musik hören"], [])], (
        "die Aenderung muss bleiben, nur der neue Skill faellt heraus"
    )
    ausgabe = capsys.readouterr().out
    assert "#1" in ausgabe, "das Issue des neuen Skills muss als abgelehnt erscheinen"
    assert "#2" not in ausgabe, "die Aenderung darf nicht abgelehnt werden"


# ── Task 2: beide Eintraege im Volltext ──────────────────────────────────────

BESTAND_ANZEIGE = {
    "hoch": {"kategorien": [{"label": "Ablenkung", "skills": [
        {"e": "🎧", "t": "Musik hören", "b": "Ein Lied auflegen und zuhoeren.",
         "tip": "💡 Kopfhoerer", "von": "Max", "erg": ""},
    ]}]},
    "mittel": {"kategorien": []},
    "tief": {"kategorien": []},
}


def test_skill_im_bestand_findet_ueber_alle_stufen():
    gefunden = vh.skill_im_bestand(BESTAND_ANZEIGE, "Musik hören")
    assert gefunden["b"] == "Ein Lied auflegen und zuhoeren."


def test_skill_im_bestand_ohne_treffer_ist_none():
    assert vh.skill_im_bestand(BESTAND_ANZEIGE, "Gibt es nicht") is None


def test_gegenueberstellung_zeigt_beide_beschreibungen():
    neu = {"emoji": "🎵", "titel": "Lieblingslied auflegen",
           "beschreibung": "Ein Lied aussuchen und nur darauf achten.",
           "tipp": "", "stufe": "Hoch", "kategorie": "Ablenkung"}
    alt = vh.skill_im_bestand(BESTAND_ANZEIGE, "Musik hören")
    t = {"titel": "Lieblingslied auflegen", "aehnlich_zu": "Musik hören",
         "stufe": "Hoch", "kategorie": "Ablenkung",
         "begruendung": "Beide beschreiben gezieltes Musikhoeren.",
         "sicherheit": "unsicher"}
    text = vh.gegenueberstellung(neu, alt, t)
    assert "Ein Lied aussuchen und nur darauf achten." in text
    assert "Ein Lied auflegen und zuhoeren." in text, "der vorhandene Text fehlt"
    assert "unsicher" in text.lower()
    assert "Beide beschreiben gezieltes Musikhoeren." in text


def test_gegenueberstellung_ohne_alten_eintrag_sagt_es():
    neu = {"emoji": "🎵", "titel": "Neu", "beschreibung": "Text.", "tipp": "",
           "stufe": "Hoch", "kategorie": "Ablenkung"}
    t = {"titel": "Neu", "aehnlich_zu": "Verschwunden", "stufe": "Hoch",
         "kategorie": "Ablenkung", "begruendung": "…", "sicherheit": "sicher"}
    text = vh.gegenueberstellung(neu, None, t)
    assert "Verschwunden" in text
    assert "nicht gefunden" in text.lower()


# ── Task 3: Drei Antworten und die Begruendung ───────────────────────────────


def uebernehmen_neu():
    return [({"number": 7}, {"art": "neu", "stufe": "Hoch", "kategorie": "Ablenkung",
              "titel": "Lieblingslied auflegen", "emoji": "🎵",
              "beschreibung": "Ein Lied aussuchen.", "tipp": ""})]


TREFFER_NEU = [{"titel": "Lieblingslied auflegen", "aehnlich_zu": "Musik hören",
                "stufe": "Hoch", "kategorie": "Ablenkung",
                "begruendung": "Beide beschreiben gezieltes Musikhoeren.",
                "sicherheit": "sicher"}]


def test_ablehnen_sammelt_nummer_und_grund():
    antworten = iter(["a", "Steht schon drin als Musik hören."])
    raus, ablehnungen = vh.nachfragen(TREFFER_NEU, uebernehmen_neu(), BESTAND_ANZEIGE,
                                      eingabe=lambda _: next(antworten))
    assert raus == [7], "abgelehnt heisst auch: nicht eintragen"
    assert ablehnungen == [(7, "Steht schon drin als Musik hören.")]


def test_ablehnen_verlangt_eine_begruendung():
    """Eine leere Begruendung waere fuer die einreichende Person wertlos."""
    antworten = iter(["a", "   ", "", "Doppelt."])
    raus, ablehnungen = vh.nachfragen(TREFFER_NEU, uebernehmen_neu(), BESTAND_ANZEIGE,
                                      eingabe=lambda _: next(antworten))
    assert ablehnungen == [(7, "Doppelt.")]


def test_uebernehmen_sammelt_keine_ablehnung():
    raus, ablehnungen = vh.nachfragen(TREFFER_NEU, uebernehmen_neu(), BESTAND_ANZEIGE,
                                      eingabe=lambda _: "ü")
    assert (raus, ablehnungen) == ([], [])


def test_weiter_sammelt_keine_ablehnung():
    raus, ablehnungen = vh.nachfragen(TREFFER_NEU, uebernehmen_neu(), BESTAND_ANZEIGE,
                                      eingabe=lambda _: "w")
    assert (raus, ablehnungen) == ([7], [])


def test_ohne_tastatur_wird_nicht_abgelehnt():
    """Ohne Eingabemoeglichkeit darf NICHTS auf GitHub geschrieben werden."""
    def keine_tastatur(_):
        raise EOFError

    raus, ablehnungen = vh.nachfragen(TREFFER_NEU, uebernehmen_neu(), BESTAND_ANZEIGE,
                                      eingabe=keine_tastatur)
    assert raus == [7]
    assert ablehnungen == [], "ohne Rueckfrage keine Ablehnung"


def test_abbruch_mitten_in_der_begruendung_lehnt_nicht_ab():
    """Endet die Eingabe mitten in der Begruendung (EOF – kein Mensch am
    Bildschirm, geschlossene Eingabe, Strg+Z), darf keine halbe Ablehnung
    zurueckbleiben: uebersprungen ja, abgelehnt nein.

    Bewusst EOF und nicht Strg+C: Ein KeyboardInterrupt nimmt einen anderen
    Weg – er wird hier nirgends gefangen und beendet den ganzen Lauf, und zwar
    noch vor `in_excel_uebernehmen`. Auch das ist sicher (nichts geschrieben,
    kein Issue veraendert), aber es ist ein anderer Fall als dieser."""
    antworten = iter(["a"])

    def dann_schluss(_):
        try:
            return next(antworten)
        except StopIteration:
            raise EOFError

    raus, ablehnungen = vh.nachfragen(TREFFER_NEU, uebernehmen_neu(), BESTAND_ANZEIGE,
                                      eingabe=dann_schluss)
    assert ablehnungen == []
    assert raus == [7]


# ── Task 4: Ablehnungen ausfuehren, automatische Faelle nach Rueckfrage ─────


def test_issue_ablehnen_kommentiert_labelt_und_schliesst(monkeypatch):
    aufrufe = []
    monkeypatch.setattr(vh.subprocess, "run",
                        lambda befehl, **k: aufrufe.append(befehl) or types.SimpleNamespace(returncode=0))
    vh.issue_ablehnen(7, "Steht schon drin.")
    flach = [" ".join(b) for b in aufrufe]
    assert any("comment" in b and "Steht schon drin." in b for b in flach)
    assert any("abgelehnt" in b for b in flach), "das Label fehlt"
    assert any("close" in b for b in flach)


def test_automatische_ablehnung_wird_nur_nach_zustimmung_kommentiert():
    faelle = [({"number": 9, "title": "Test"}, "Kategorie existiert nicht")]
    assert vh.automatische_ablehnungen_melden(faelle, eingabe=lambda _: "n") == []
    zu_schreiben = vh.automatische_ablehnungen_melden(faelle, eingabe=lambda _: "j")
    assert len(zu_schreiben) == 1
    assert zu_schreiben[0][0] == 9
    assert "Kategorie existiert nicht" in zu_schreiben[0][1]


def test_automatische_ablehnung_ohne_tastatur_schreibt_nichts():
    def keine_tastatur(_):
        raise EOFError

    faelle = [({"number": 9, "title": "Test"}, "Kategorie existiert nicht")]
    assert vh.automatische_ablehnungen_melden(faelle, eingabe=keine_tastatur) == []


def test_main_lehnt_erst_nach_dem_schreiben_ab(tmp_path, monkeypatch):
    """Die Reihenfolge ist die Zusicherung: erst die Mappe, dann GitHub."""
    reihenfolge = []
    aufrufe = []
    antworten = iter(["a", "Steht schon drin."])
    # Genau ein Treffer, der Mensch lehnt ihn beim Nachfragen explizit ab
    # (mit Begruendung) – das fuellt vh.main()s `ablehnungen`-Liste mit genau
    # einem Eintrag, der danach ueber issue_ablehnen() geschrieben wird.
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: [{
            "titel": "Spazieren gehen", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: next(antworten),
    )
    # Die drei entscheidenden Schreibvorgaenge werden erst hier ueberschrieben
    # (nach _duplikat_lauf, das issue_schliessen/in_excel_uebernehmen bereits
    # setzt) – nur so landen ihre Aufrufe in derselben Reihenfolge-Liste.
    monkeypatch.setattr(vh, "in_excel_uebernehmen",
                        lambda *a, **k: reihenfolge.append("excel") or 1)
    monkeypatch.setattr(vh, "issue_ablehnen",
                        lambda *a, **k: reihenfolge.append("ablehnen"))
    monkeypatch.setattr(vh, "issue_schliessen",
                        lambda *a, **k: reihenfolge.append("schliessen"))

    vh.main()

    assert reihenfolge.index("excel") < reihenfolge.index("ablehnen")


def test_fehler_beim_ablehnen_bricht_den_lauf_nicht_ab(capsys, monkeypatch):
    """Die Uebernahme steht hier schon in der Mappe – ein Abbruch waere fatal."""
    aufrufe = []
    antworten = iter(["a", "Steht schon drin."])
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: [{
            "titel": "Spazieren gehen", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: next(antworten),
    )

    def schlaegt_wie_gh_fehl(nummer, grund):
        raise subprocess.CalledProcessError(1, ["gh"])

    monkeypatch.setattr(vh, "issue_ablehnen", schlaegt_wie_gh_fehl)

    vh.main()  # darf NICHT platzen

    ausgabe = capsys.readouterr().out
    assert "von Hand" in ausgabe


def test_automatischer_fall_wird_nicht_ein_zweites_mal_erfragt(monkeypatch):
    """`abgelehnt` enthaelt auch die bei der Duplikat-Rueckfrage uebersprungenen
    Issues (ueber `raus`). Wuerde main() die ganze Liste an
    automatische_ablehnungen_melden() weiterreichen, bekaeme der Mensch fuer
    genau das Issue, das er soeben mit „weiter" beantwortet hat, eine zweite
    Rueckfrage."""
    aufrufe = []
    gefragt = []
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: [{
            "titel": "Spazieren gehen", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: "w",
    )

    echte = vh.automatische_ablehnungen_melden

    def aufgezeichnet(faelle, eingabe=input):
        gefragt.append([i["number"] for i, _ in faelle])
        return echte(faelle, eingabe=eingabe)

    monkeypatch.setattr(vh, "automatische_ablehnungen_melden", aufgezeichnet)

    vh.main()

    assert gefragt == [[]], "Issue #1 wurde beim Nachfragen schon beantwortet"


# ── Fix-Runde nach Pruefung: automatischer Pfad, Widerspruch, Meldungstext ──


def test_automatischer_pfad_schliesst_und_labelt_niemals(monkeypatch):
    """Der Docstring von automatische_ablehnungen_melden verspricht: diese
    Faelle werden NICHT geschlossen – sie sind oft behebbar (Kategorie zuerst
    anlegen) und sollen beim naechsten Lauf wieder angeboten werden. Der
    automatische Pfad in main() darf deshalb NUR kommentieren, niemals
    schliessen oder labeln."""
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, {**BEISPIEL, "kategorie": "Erfunden"}),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 0)
    monkeypatch.setattr(
        vh, "automatische_ablehnungen_melden",
        lambda faelle: [(i, f"Nicht uebernommen: {g}") for i, g in [(1, "x")]],
    )
    aufrufe = []
    monkeypatch.setattr(vh.subprocess, "run",
                        lambda befehl, **k: aufrufe.append(befehl) or types.SimpleNamespace(returncode=0))

    vh.main()

    flach = [" ".join(b) for b in aufrufe]
    assert any("comment" in b for b in flach), "der Kommentar muss geschrieben werden"
    assert not any("close" in b for b in flach), (
        "der automatische Pfad darf das Issue nicht schliessen"
    )
    assert not any("--add-label" in b for b in flach), (
        "der automatische Pfad darf kein Label setzen"
    )


def test_erfolgreich_abgelehntes_issue_erscheint_nicht_als_bleibt_offen(monkeypatch, capsys):
    """Ein per `a` abgelehntes und erfolgreich geschlossenes Issue darf nicht
    zusaetzlich als "bleibt offen" auftauchen – das waere eine falsche Aussage
    ueber einen Vorgang, den das Programm selbst gerade ausgefuehrt hat."""
    aufrufe = []
    antworten = iter(["a", "Steht schon drin."])
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: [{
            "titel": "Spazieren gehen", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: next(antworten),
    )
    # issue_ablehnen NICHT gemockt: die echte Funktion laeuft ueber das in
    # _duplikat_lauf gemockte, immer erfolgreiche subprocess.run durch.

    vh.main()

    ausgabe = capsys.readouterr().out
    assert "abgelehnt und geschlossen" in ausgabe
    assert "bleibt offen" not in ausgabe


def test_bleibt_offen_erscheint_nur_fuer_tatsaechlich_offene_issues(monkeypatch, capsys):
    """Zwei Faelle im selben Lauf: Issue #1 wird per `a` abgelehnt und
    erfolgreich geschlossen, Issue #2 scheitert automatisch an der Pruefung
    und bleibt tatsaechlich offen. Nur #2 darf als "bleibt offen" auftauchen,
    und der Schlusstext ("Was tun?") muss weiterhin erscheinen, weil #2 noch
    offen ist."""
    aufrufe = []
    antworten = iter(["a", "Steht schon drin."])
    _duplikat_lauf(
        monkeypatch, [
            freigegebenes_issue(1, NEUER_SKILL),
            freigegebenes_issue(2, {**BEISPIEL, "kategorie": "Erfunden"}),
        ],
        lambda neue, bestand, client: [{
            "titel": "Spazieren gehen", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: next(antworten),
    )
    monkeypatch.setattr(vh, "automatische_ablehnungen_melden", lambda faelle: [])

    vh.main()

    ausgabe = capsys.readouterr().out
    assert "abgelehnt und geschlossen" in ausgabe
    zeilen_mit_1 = [z for z in ausgabe.splitlines() if "#1" in z]
    assert not any("bleibt offen" in z for z in zeilen_mit_1), (
        "Issue #1 wurde gerade geschlossen, darf nicht als offen erscheinen"
    )
    zeilen_mit_2 = [z for z in ausgabe.splitlines() if "#2" in z]
    assert any("bleibt offen" in z for z in zeilen_mit_2), (
        "Issue #2 ist tatsaechlich noch offen und muss das auch sagen"
    )
    assert "Was tun?" in ausgabe, "Issue #2 ist noch offen, der Schlusstext muss erscheinen"


# ── Abschlusspruefung B-1: der Wachposten fuer den Fall NACH dem Schreiben ──
#
# Zwischen der Schliess-Schleife (mit `except BaseException` abgesichert) und
# der abschliessenden Warnung stehen seit diesem Zweig zwei neue Bloecke: die
# `ablehnungen`-Schleife und `automatische_ablehnungen_melden` – Letztere mit
# einem interaktiven `input()`. Bricht dort etwas ab, steht die Uebernahme
# bereits in der Mappe. Ohne Warnung bleibt unbemerkt, dass Issues offen sind,
# und der naechste Lauf traegt dieselben Skills ein zweites Mal ein.


def test_abbruch_bei_der_zweiten_rueckfrage_warnt_vor_offenen_issues(monkeypatch, capsys):
    """Strg+C an der Rueckfrage zu den aussortierten Vorschlaegen: die Mappe ist
    geschrieben, Issue #1 liess sich nicht schliessen. Die Warnung MUSS
    erscheinen, sonst erzeugt der naechste Lauf eine Dublette."""
    monkeypatch.setattr(vh, "hole_issues", lambda: [
        freigegebenes_issue(1, BEISPIEL),
        freigegebenes_issue(2, {**BEISPIEL, "kategorie": "Erfunden"}),
    ])
    monkeypatch.setattr(vh, "lade_datenstand", lambda: BESTAND)
    monkeypatch.setattr(vh, "in_excel_uebernehmen", lambda pfad, aenderungen, neue: 1)
    # Die Duplikatpruefung ist hier nicht das Thema – und ohne diese Attrappe
    # haenge das Ergebnis davon ab, ob auf dem Rechner ein Schluessel liegt.
    monkeypatch.setattr(vh.duplikat, "schluessel_finden", lambda projekt: None)

    def schlaegt_wie_gh_fehl(nummer):
        raise subprocess.CalledProcessError(1, ["gh"])

    monkeypatch.setattr(vh, "issue_schliessen", schlaegt_wie_gh_fehl)

    def strg_c(_):
        raise KeyboardInterrupt

    echte = vh.automatische_ablehnungen_melden
    monkeypatch.setattr(
        vh, "automatische_ablehnungen_melden",
        lambda faelle: echte(faelle, eingabe=strg_c),
    )

    with pytest.raises(KeyboardInterrupt):
        vh.main()

    ausgabe = capsys.readouterr().out
    assert "konnte(n) nicht geschlossen werden" in ausgabe
    assert "#1" in ausgabe
    assert "bereits in der Excel" in ausgabe
    # Ohne diese Zeile weiss niemand, dass ueberhaupt etwas geschrieben wurde –
    # der Lauf endet sonst im rohen Traceback.
    assert "✅" in ausgabe


def test_unerwarteter_fehler_beim_ablehnen_warnt_vor_offenen_issues(monkeypatch, capsys):
    """Dieselbe Luecke eine Stufe frueher: ein Programmierfehler in der
    `ablehnungen`-Schleife (kein CalledProcessError) schlaegt durch. Auch dann
    steht die Uebernahme schon in der Mappe."""
    aufrufe = []
    antworten = iter(["a", "Steht schon drin."])
    _duplikat_lauf(
        monkeypatch, [
            freigegebenes_issue(1, NEUER_SKILL),
            freigegebenes_issue(2, {**NEUER_SKILL, "titel": "Zweiter Skill"}),
        ],
        lambda neue, bestand, client: [{
            "titel": "Zweiter Skill", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: next(antworten),
    )

    def schliessen_scheitert(nummer):
        raise subprocess.CalledProcessError(1, ["gh"])

    monkeypatch.setattr(vh, "issue_schliessen", schliessen_scheitert)

    def schlaegt_unerwartet_fehl(nummer, grund):
        raise RuntimeError("Programmierfehler")

    monkeypatch.setattr(vh, "issue_ablehnen", schlaegt_unerwartet_fehl)

    with pytest.raises(RuntimeError):
        vh.main()

    ausgabe = capsys.readouterr().out
    assert "konnte(n) nicht geschlossen werden" in ausgabe
    assert "#1" in ausgabe


def test_meldung_nach_gescheitertem_ablehnen_nennt_den_schon_geschriebenen_kommentar(
    monkeypatch, capsys
):
    """`issue_ablehnen` kommentiert zuerst und labelt danach. Scheitert es erst
    dort, steht die Begruendung bereits oeffentlich im Issue. Die Meldung darf
    den Menschen dann nicht dazu auffordern, sie ein zweites Mal zu schreiben."""
    aufrufe = []
    antworten = iter(["a", "Steht schon drin."])
    _duplikat_lauf(
        monkeypatch, [freigegebenes_issue(1, NEUER_SKILL)],
        lambda neue, bestand, client: [{
            "titel": "Spazieren gehen", "aehnlich_zu": "Musik hören",
            "stufe": "Hoch", "kategorie": "Ablenkung",
            "begruendung": "Dieselbe Handlung.", "sicherheit": "unsicher",
        }],
        aufrufe,
        eingabe=lambda _: next(antworten),
    )

    def schlaegt_wie_gh_fehl(nummer, grund):
        raise subprocess.CalledProcessError(1, ["gh"])

    monkeypatch.setattr(vh, "issue_ablehnen", schlaegt_wie_gh_fehl)

    vh.main()

    ausgabe = capsys.readouterr().out
    assert "konnte nicht abgelehnt werden" in ausgabe
    # Nur die Meldung selbst betrachten: der Schlusstext („Was tun?") spricht
    # ebenfalls von Kommentaren und wuerde eine unscharfe Pruefung bestehen.
    meldung = ausgabe.split("konnte nicht abgelehnt werden")[1].split("⚠")[0]
    assert "von Hand" in meldung
    assert "moeglicherweise schon" in meldung, (
        "die Meldung muss sagen, dass der Kommentar moeglicherweise schon steht"
    )
