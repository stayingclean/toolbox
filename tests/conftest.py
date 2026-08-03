"""Gemeinsame Testhilfen: Repo-Wurzel importierbar machen + Minimal-Excel bauen."""

import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

STUFEN_HEADER = ["Stufe", "Bezeichnung", "Bereich", "Icon", "Intro",
                 "Farbe", "Farbe2", "Hell"]
STUFEN_ROWS = [
    ["Hoch", "Hohe Anspannung", "80-100", "🌶️", "Intro hoch", "#a00", "#c00", "#fee"],
    ["Mittel", "Mittlere Anspannung", "40-79", "🌤️", "Intro mittel", "#0a0", "#0c0", "#efe"],
    ["Tief", "Tiefe Anspannung", "0-39", "🌊", "Intro tief", "#00a", "#00c", "#eef"],
]
KATEGORIEN_HEADER = ["Stufe", "Kategorie", "Icon"]
KATEGORIEN_ROWS = [["Hoch", "Ablenkung", "🎧"]]


@pytest.fixture
def mappe(tmp_path):
    """Erzeugt eine gültige skills_daten.xlsx mit frei wählbarem Skills-Blatt."""

    def bauen(skills_header, skills_rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Skills"
        ws.append(skills_header)
        for zeile in skills_rows:
            ws.append(zeile)
        ws2 = wb.create_sheet("Stufen")
        ws2.append(STUFEN_HEADER)
        for zeile in STUFEN_ROWS:
            ws2.append(zeile)
        ws3 = wb.create_sheet("Kategorien")
        ws3.append(KATEGORIEN_HEADER)
        for zeile in KATEGORIEN_ROWS:
            ws3.append(zeile)
        pfad = tmp_path / "skills_daten.xlsx"
        wb.save(pfad)
        return pfad

    return bauen
