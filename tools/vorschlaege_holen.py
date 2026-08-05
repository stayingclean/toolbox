# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Freigegebene Skill-Vorschläge übernehmen
========================================

Holt aus dem Vorschlags-Repo alle offenen Issues mit dem Label „freigegeben",
überträgt sie ins Blatt `Skills` von skills_daten.xlsx, schliesst die Issues und
ruft anschliessend build.py auf.

Zwei Arten von Vorschlägen:
  • `neu`       – wird als neue Zeile angehängt
  • `aenderung` – ersetzt die Texte der bestehenden Zeile (gefunden über Stufe +
                  Kategorie + ursprünglicher Titel). `Von` bleibt stehen, die
                  ergänzende Person kommt in `Ergaenzt` dazu.

Aufruf:  vorschlaege.bat      (oder `uv run tools/vorschlaege_holen.py`)

Der Push bleibt bewusst von Hand — nach dem Lauf zuerst die Skillsliste
anschauen, dann committen und pushen.

Voraussetzung: das GitHub-CLI `gh` ist installiert und angemeldet
(`gh auth status` muss ein angemeldetes Konto zeigen).
"""

import json
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "skills_daten.xlsx"
BUILD = ROOT / "build.py"
DATEN_JSON = ROOT / "docs" / "skills-daten.json"
REPO = "stayingclean/skills-suggestions"
LABEL = "freigegeben"

# Nur Issues von diesem Konto werden uebernommen. Der Worker legt sie unter dem
# Token dieses Kontos an; von Hand eroeffnete Issues sind nie geprueft worden.
BOT = "eraschle"

# Muessen zu den Grenzen in worker/validate.js passen.
GRENZEN = {
    "titel": 60,
    "beschreibung": 300,
    "tipp": 200,
    "von": 30,
}

# Der Worker zaehlt beim Emoji Graphem-Cluster (sichtbare Zeichen) ueber
# Intl.Segmenter, damit ein zusammengesetztes Emoji wie 🧘‍♀️ (vier Codepunkte,
# aber EIN sichtbares Zeichen) nicht faelschlich abgelehnt wird. Python hat
# keine eingebaute Graphem-Zerlegung, und eine Abhaengigkeit dafuer wollen wir
# nicht – darum hier bewusst nur eine grobe, grosszuegige Schranke in
# Codepunkten: ein einzelnes Emoji bleibt weit darunter, aber beliebig lange
# Ketten werden trotzdem abgelehnt.
EMOJI_CODEPUNKT_GRENZE = 16
FELDNAMEN = {
    "emoji": "Emoji",
    "titel": "Titel",
    "beschreibung": "Beschreibung",
    "tipp": "Tipp",
    "von": "Name",
}
PFLICHT = ["stufe", "kategorie", "emoji", "titel", "beschreibung"]
TEXTFELDER = ["titel", "beschreibung", "tipp", "von"]
# Anzeigename -> Schluessel in docs/skills-daten.json
STUFEN = {"Hoch": "hoch", "Mittel": "mittel", "Tief": "tief"}

SPALTEN = [
    ("Stufe", "stufe"),
    ("Kategorie", "kategorie"),
    ("Emoji", "emoji"),
    ("Titel", "titel"),
    ("Beschreibung", "beschreibung"),
    ("Tipp", "tipp"),
    ("Von", "von"),
]

# Bei einer Aenderung werden NUR diese Spalten ueberschrieben. `Stufe` und
# `Kategorie` bleiben, weil sie mit dem Titel den Schluessel zur alten Zeile
# bilden; `Von` bleibt, damit die urspruenglich beitragende Person genannt bleibt.
SPALTEN_AENDERUNG = [
    ("Emoji", "emoji"),
    ("Titel", "titel"),
    ("Beschreibung", "beschreibung"),
    ("Tipp", "tipp"),
    ("Ergaenzt", "erg"),
]

# Ueber diese drei Spalten wird die zu aendernde Zeile gesucht.
SCHLUESSEL_SPALTEN = ["Stufe", "Kategorie", "Titel"]


class ZeileNichtGefunden(Exception):
    """Der zu aendernde Skill steht nicht (mehr) in der Excel."""


class ZeileMehrdeutig(Exception):
    """Mehrere Zeilen passen auf denselben Schluessel – hier wird nicht geraten."""


BLOCK = re.compile(r"<!--\s*vorschlag\s*(\{.*?\})\s*-->", re.DOTALL)


def parse_body(body: str):
    """Liest den maschinenlesbaren Block aus einem Issue-Rumpf.

    Liefert None, wenn kein Block vorhanden ist, das JSON nicht lesbar ist oder
    MEHR ALS EIN Block gefunden wird. Der letzte Fall ist der Abwehrschritt gegen
    einen gefaelschten Block, den jemand in ein Freitextfeld geschrieben hat –
    solche Issues bleiben offen und werden gemeldet.
    """
    treffer = BLOCK.findall(body or "")
    if len(treffer) != 1:
        return None
    try:
        daten = json.loads(treffer[0])
    except json.JSONDecodeError:
        return None
    return daten if isinstance(daten, dict) else None


def pruefe_eintrag(eintrag: dict, daten: dict):
    """Prueft einen Vorschlag gegen den aktuellen Datenbestand.

    Liefert None, wenn alles stimmt, sonst eine verstaendliche Meldung.
    Der Worker prueft dasselbe – hier geht es um Issues, die nicht ueber das
    Formular kamen, und um Vorschlaege, deren Kategorie inzwischen weg ist.
    """
    art = str(eintrag.get("art") or "neu").strip()
    if art not in ("neu", "aenderung"):
        return f"Unbekannte Art: {art!r}."

    for schluessel in PFLICHT:
        wert = eintrag.get(schluessel)
        if not isinstance(wert, str) or not wert.strip():
            name = FELDNAMEN.get(schluessel, schluessel.capitalize())
            return f"Pflichtfeld fehlt oder ist leer: {name}."

    # Tipp und Von duerfen fehlen und gelten dann als leer.
    felder = {s: str(eintrag.get(s) or "").strip() for s in GRENZEN}
    felder["emoji"] = str(eintrag.get("emoji") or "").strip()
    felder.update({s: str(eintrag.get(s) or "").strip() for s in ("stufe", "kategorie")})

    # Bei einer Aenderung traegt `erg` den Namen. Er nimmt hier den Platz von
    # `von` ein, damit die Laengen-, Link- und Zeichenpruefungen darunter
    # unveraendert gelten – die Meldung nennt in beiden Faellen „Name".
    if art == "aenderung":
        felder["von"] = str(eintrag.get("erg") or "").strip()

    # Emoji steht nicht in GRENZEN: eigenstaendige, groebere Pruefung (siehe
    # Kommentar bei EMOJI_CODEPUNKT_GRENZE) an der Stelle, an der zuvor die
    # Laengenpruefung fuer Emoji lief.
    if len(felder["emoji"]) > EMOJI_CODEPUNKT_GRENZE:
        return "Emoji ist zu lang."

    for schluessel, grenze in GRENZEN.items():
        if len(felder[schluessel]) > grenze:
            return (
                f"Zu lang: {FELDNAMEN[schluessel]} "
                f"({len(felder[schluessel])} statt hoechstens {grenze} Zeichen)."
            )

    for schluessel in TEXTFELDER:
        if "http" in felder[schluessel].lower():
            return "Links sind nicht erlaubt."

    for schluessel in TEXTFELDER + ["emoji"]:
        wert = felder[schluessel]
        if "<!--" in wert or "-->" in wert:
            return "Kommentarzeichen sind nicht erlaubt."
        if "<" in wert or ">" in wert:
            return "Spitze Klammern sind nicht erlaubt."

    stufe = felder["stufe"]
    if stufe not in STUFEN:
        return f"Unbekannte Stufe: '{stufe}' (erlaubt sind Hoch, Mittel, Tief)."

    stufen_daten = daten.get(STUFEN[stufe]) or {}
    kategorien = [k.get("label") for k in stufen_daten.get("kategorien", [])]
    if felder["kategorie"] not in kategorien:
        return (
            f"Unbekannte Kategorie: '{felder['kategorie']}' "
            f"gibt es in der Stufe '{stufe}' nicht (mehr)."
        )

    if art == "aenderung":
        original = str(eintrag.get("original") or "").strip()
        if not original:
            return "Pflichtfeld fehlt: urspruenglicher Titel."
        kat = next(
            (
                k
                for k in stufen_daten.get("kategorien", [])
                if k.get("label") == felder["kategorie"]
            ),
            None,
        )
        if not any(s.get("t") == original for s in (kat or {}).get("skills", [])):
            return (
                f"Der Skill '{original}' steht nicht mehr in der Stufe "
                f"'{stufe}', Kategorie '{felder['kategorie']}' – vermutlich "
                f"inzwischen umbenannt oder entfernt."
            )

    return None


def bereinigt(eintrag: dict) -> dict:
    """Liefert eine Kopie mit getrimmten Werten fuer alle Excel-Spalten.

    pruefe_eintrag() prueft Laengen und Inhalte auf der getrimmten Fassung
    (`.strip()`), aber der rohe Eintrag enthaelt noch den Randweissraum. Ohne
    diesen Schritt wuerde z. B. ein Titel aus 60 Zeichen plus zwei Leerzeichen
    die Pruefung bestehen und trotzdem mit 62 Zeichen in der Excel landen.
    """
    ergebnis = dict(eintrag)
    for schluessel in {s for _, s in SPALTEN} | {s for _, s in SPALTEN_AENDERUNG}:
        ergebnis[schluessel] = str(eintrag.get(schluessel) or "").strip()
    return ergebnis


def lade_datenstand() -> dict:
    """Liest docs/skills-daten.json – den Stand, gegen den geprueft wird."""
    if not DATEN_JSON.exists():
        raise SystemExit(
            f"❌ {DATEN_JSON.name} fehlt.\n\n"
            "   Ohne diese Datei laesst sich nicht pruefen, welche Stufen und\n"
            "   Kategorien es gibt. Bitte zuerst build.bat ausfuehren."
        )
    try:
        return json.loads(DATEN_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        raise SystemExit(
            f"❌ {DATEN_JSON.name} ist nicht lesbar.\n\n"
            "   Bitte zuerst build.bat ausfuehren."
        )


def hole_issues():
    """Fragt alle offenen Issues über das GitHub-CLI ab (ungefiltert).

    Der serverseitige Label-Filter von GitHub (`--label`) ist nicht sofort
    aktuell: ein frisch gesetztes Label taucht dort erst nach einigen Sekunden
    auf. Darum werden hier alle offenen Issues samt Labels geholt und das
    Filtern nach `hat_label()` in Python erledigt.
    """
    try:
        ergebnis = subprocess.run(
            ["gh", "issue", "list", "--repo", REPO,
             "--state", "open", "--limit", "100",
             "--json", "number,title,body,author,labels"],
            capture_output=True, text=True, encoding="utf-8",
        )
    except FileNotFoundError:
        raise SystemExit(
            "❌ Das GitHub-Programm `gh` wurde nicht gefunden.\n\n"
            "   Lade es von https://cli.github.com herunter, installiere es und\n"
            "   melde dich danach einmal mit  gh auth login  an.\n"
            "   Ohne `gh` kann dieses Skript die Vorschlaege nicht abholen."
        )
    if ergebnis.returncode != 0:
        raise SystemExit(
            "❌ Konnte die Vorschläge nicht abrufen:\n"
            + ergebnis.stderr.strip()
            + "\n\nIst `gh` installiert und angemeldet? Prüfe mit: gh auth status"
        )
    return json.loads(ergebnis.stdout or "[]")


def hat_label(issue: dict, name: str) -> bool:
    """Prueft das Label in Python statt ueber den Server.

    Der serverseitige Label-Filter von GitHub ist nicht sofort aktuell: ein
    frisch gesetztes Label taucht dort erst nach einigen Sekunden auf. Wer
    freigibt und sofort startet, bekaeme sonst faelschlich „nichts zu tun".
    """
    return any(l.get("name") == name for l in issue.get("labels", []))


def speichern(wb, pfad: Path):
    """Speichert die Mappe – oder sagt verstaendlich, warum es nicht geht.

    Geschrieben wird zuerst in eine Nachbardatei; erst wenn sie vollstaendig
    dasteht, wird sie an die Stelle der alten gelegt (`os.replace`, unter
    Windows ein einziger Schritt). Bricht das Schreiben mittendrin ab – volle
    Festplatte, abgezogener USB-Stick –, bleibt die alte Datei unversehrt.
    Direkt auf die Zieldatei geschrieben waere sie in diesem Moment auf halber
    Laenge abgeschnitten und nicht mehr lesbar; die Excel ist aber das
    Original, aus dem alles andere erzeugt wird, und es gibt keine zweite
    Quelle.

    Es gibt genau EINEN Speichervorgang je Lauf (siehe in_excel_uebernehmen),
    und er ist der letzte Schritt. Nur darum darf die Meldung versprechen, dass
    nichts veraendert wurde.
    """
    zwischendatei = pfad.with_name(pfad.name + ".neu")
    try:
        wb.save(zwischendatei)
        os.replace(zwischendatei, pfad)
    except PermissionError:
        raise SystemExit(
            f"❌ Die Datei {pfad.name} laesst sich nicht speichern.\n\n"
            f"   Sie ist vermutlich gerade in Excel geoeffnet. Bitte schliesse\n"
            f"   Excel und starte vorschlaege.bat noch einmal.\n\n"
            f"   Es wurde nichts veraendert – die Vorschlaege bleiben freigegeben\n"
            f"   und werden beim naechsten Lauf uebernommen."
        )
    except OSError as fehler:
        raise SystemExit(
            f"❌ Die Datei {pfad.name} konnte nicht geschrieben werden.\n\n"
            f"   Der Computer meldet: {fehler.strerror or fehler}\n\n"
            f"   Meist ist die Festplatte voll, oder die Datei liegt auf einem\n"
            f"   Laufwerk (USB-Stick, Netzlaufwerk), das gerade nicht erreichbar\n"
            f"   ist. Bitte Platz schaffen bzw. das Laufwerk pruefen und\n"
            f"   vorschlaege.bat noch einmal starten.\n\n"
            f"   Die alte Fassung von {pfad.name} ist unveraendert erhalten\n"
            f"   geblieben – die Vorschlaege bleiben freigegeben."
        )
    finally:
        # Nach dem Umlegen ist die Nachbardatei weg; nach einem Fehlschlag darf
        # sie nicht liegen bleiben und jemanden verwirren.
        try:
            zwischendatei.unlink(missing_ok=True)
        except OSError:
            pass


def spalten_sichern(ws, kopf: list, spalten: list) -> bool:
    """Legt fehlende Spalten rechts an. True, wenn eine dazugekommen ist."""
    dazugekommen = False
    for name, _ in spalten:
        if name not in kopf:
            ws.cell(row=1, column=len(kopf) + 1, value=name)
            kopf.append(name)
            dazugekommen = True
    return dazugekommen


def zeile_ersetzen(ws, kopf: list, eintrag: dict):
    """Sucht die Zeile eines bestehenden Skills und ersetzt ihre Texte.

    Gefunden wird ueber Stufe + Kategorie + urspruenglicher Titel. Die Spalte
    `Von` bleibt unangetastet – der urspruengliche Beitragende wird nie durch
    eine Ergaenzung verdraengt; die ergaenzende Person kommt in `Ergaenzt` dazu.
    """
    e = bereinigt(eintrag)
    original = str(eintrag.get("original") or "").strip()
    if not original:
        # Ohne diese Sperre wuerde ein leerer Titel auf eine Zeile mit leerer
        # Titelzelle passen und die falsche Zeile ueberschreiben. pruefe_eintrag
        # faengt den Fall ebenfalls ab – hier steht er noch einmal, direkt vor
        # dem Schreiben, weil die Excel nicht wiederherstellbar ist.
        raise ZeileNichtGefunden(
            f"ohne urspruenglichen Titel ({e['stufe']} / {e['kategorie']})"
        )

    i_stufe, i_kat, i_titel = (kopf.index(n) + 1 for n in SCHLUESSEL_SPALTEN)

    def zelle(zeile, spalte):
        wert = ws.cell(row=zeile, column=spalte).value
        return str(wert).strip() if wert is not None else ""

    treffer = [
        r
        for r in range(2, ws.max_row + 1)
        if zelle(r, i_stufe) == e["stufe"]
        and zelle(r, i_kat) == e["kategorie"]
        and zelle(r, i_titel) == original
    ]
    if not treffer:
        raise ZeileNichtGefunden(f"'{original}' in {e['stufe']} / {e['kategorie']}")
    if len(treffer) > 1:
        # Nicht die erste Zeile nehmen: dann wuerde womoeglich der falsche
        # Eintrag geaendert, ohne dass es jemand merkt.
        zeilen = ", ".join(str(r) for r in treffer)
        raise ZeileMehrdeutig(
            f"'{original}' in {e['stufe']} / {e['kategorie']} "
            f"steht in den Zeilen {zeilen}"
        )
    for name, schluessel in SPALTEN_AENDERUNG:
        ws.cell(row=treffer[0], column=kopf.index(name) + 1, value=e.get(schluessel, ""))


def in_excel_uebernehmen(pfad: Path, aenderungen: list, neue: list) -> int:
    """Traegt Aenderungen und neue Skills ins Blatt `Skills` ein.

    Bewusst EIN Laden-Aendern-Speichern fuer den ganzen Lauf. Mit zwei
    Durchgaengen koennte der zweite scheitern, nachdem der erste bereits
    gespeichert hat: dann staende die halbe Uebernahme in der Excel, waehrend
    alle Issues noch offen sind – und der naechste Lauf faende die schon
    geaenderte Zeile nicht mehr wieder. So gibt es nur einen Speichervorgang,
    der gelingt oder scheitert; bei jedem Abbruch bleibt die Datei unberuehrt.

    Erst die Aenderungen, dann die neuen Zeilen: eine Aenderung darf sich nie
    auf einen Skill beziehen, den derselbe Lauf erst angelegt hat.

    Liefert die Anzahl der geschriebenen Zeilen.
    """
    if not aenderungen and not neue:
        return 0
    if not pfad.exists():
        raise SystemExit(
            f"❌ Die Datei {pfad.name} wurde nicht gefunden.\n\n"
            f"   Erwartet wird sie hier:\n"
            f"   {pfad}\n\n"
            f"   Wurde sie verschoben oder umbenannt? Bitte zuruecklegen und\n"
            f"   vorschlaege.bat noch einmal starten.\n\n"
            f"   Es wurde nichts veraendert – die Vorschlaege bleiben freigegeben."
        )
    try:
        wb = openpyxl.load_workbook(pfad)
    except (OSError, ValueError, zipfile.BadZipFile, InvalidFileException) as fehler:
        raise SystemExit(
            f"❌ Die Datei {pfad.name} laesst sich nicht lesen.\n\n"
            f"   Vermutlich ist sie beschaedigt. Oeffne sie einmal in Excel und\n"
            f"   speichere sie neu – oder hole die letzte funktionierende Fassung\n"
            f"   zurueck (in GitHub Desktop: Rechtsklick auf die Datei unter\n"
            f"   'Changes' und 'Discard changes').\n\n"
            f"   Es wurde nichts veraendert – die Vorschlaege bleiben freigegeben.\n\n"
            f"   (Technische Meldung: {fehler})"
        )
    if "Skills" not in wb.sheetnames:
        vorhanden = ", ".join(wb.sheetnames) or "keine"
        raise SystemExit(
            f"❌ In {pfad.name} gibt es kein Blatt namens 'Skills'.\n\n"
            f"   Vorhandene Blaetter: {vorhanden}\n\n"
            f"   Das Blatt wurde vermutlich umbenannt. Bitte den Namen 'Skills'\n"
            f"   wiederherstellen und vorschlaege.bat noch einmal starten.\n\n"
            f"   Es wurde nichts veraendert – die Vorschlaege bleiben freigegeben."
        )
    ws = wb["Skills"]
    kopf = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    neue_spalte = False
    if aenderungen:
        fehlend = [name for name in SCHLUESSEL_SPALTEN if name not in kopf]
        if fehlend:
            raise SystemExit(
                f"❌ Im Blatt 'Skills' von {pfad.name} fehlen die Spalten: "
                f"{', '.join(fehlend)}.\n\n"
                f"   Ohne sie laesst sich nicht finden, welche Zeile geaendert werden\n"
                f"   soll. Bitte die Kopfzeile nicht umbenennen.\n\n"
                f"   Es wurde nichts veraendert – die Vorschlaege bleiben freigegeben."
            )
        neue_spalte = spalten_sichern(ws, kopf, SPALTEN_AENDERUNG)
        for eintrag in aenderungen:
            zeile_ersetzen(ws, kopf, eintrag)

    if neue:
        neue_spalte = spalten_sichern(ws, kopf, SPALTEN) or neue_spalte
        for eintrag in neue:
            zeile = [""] * len(kopf)
            for name, schluessel in SPALTEN:
                zeile[kopf.index(name)] = eintrag.get(schluessel, "")
            ws.append(zeile)
        # Filterbereich und Stufen-Dropdown auf die neuen Zeilen ausdehnen, sonst
        # fallen sie heraus und beim Sortieren koennen Werte verrutschen.
        letzte = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(kopf))}{letzte}"
        for pruefung in ws.data_validations.dataValidation:
            if str(pruefung.sqref).startswith("A2:A"):
                pruefung.sqref = f"A2:A{letzte}"
    elif neue_spalte and ws.auto_filter.ref:
        # Ohne neue Zeilen bleibt der Filter, wie er ist – ausser es ist eine
        # Spalte dazugekommen, dann muss er breiter werden.
        ws.auto_filter.ref = f"A1:{get_column_letter(len(kopf))}{ws.max_row}"

    speichern(wb, pfad)
    return len(aenderungen) + len(neue)


def issue_schliessen(nummer: int):
    subprocess.run(
        ["gh", "issue", "close", str(nummer), "--repo", REPO,
         "--comment", "Übernommen – erscheint mit dem nächsten Build in der Skillsliste."],
        check=True, capture_output=True, text=True, encoding="utf-8",
    )


def warne_offene_issues(nummern: list):
    """Warnt vor Vorschlaegen, die in der Excel stehen, deren Issue aber offen ist."""
    if not nummern:
        return
    liste = ", ".join(f"#{n}" for n in nummern)
    print(
        f"\n⚠ ACHTUNG: {liste} konnte(n) nicht geschlossen werden.\n"
        f"   Die Vorschlaege sind bereits in der Excel. Schliesse diese Issues\n"
        f"   von Hand auf github.com/{REPO}/issues, sonst versucht der naechste\n"
        f"   Lauf sie noch einmal zu uebernehmen: ein neuer Skill stuende dann\n"
        f"   doppelt in der Liste, und eine Aenderung liesse sich nicht mehr\n"
        f"   zuordnen."
    )


def main():
    alle_offenen = hole_issues()
    issues = [i for i in alle_offenen if hat_label(i, LABEL)]
    if not issues:
        print("Keine freigegebenen Vorschläge offen. Nichts zu tun.")
        if alle_offenen:
            print(f"({len(alle_offenen)} Vorschlaege warten noch auf deine Freigabe.)")
        return

    bestand = lade_datenstand()

    uebernehmen, uebersprungen, abgelehnt = [], [], []
    for issue in issues:
        # Herkunft zuerst: das Repo ist oeffentlich, jede Person mit GitHub-Konto
        # kann dort ein Issue eroeffnen – solche Rumpfe hat nie jemand geprueft.
        konto = (issue.get("author") or {}).get("login", "")
        if konto != BOT:
            abgelehnt.append(
                (issue, f"stammt von '{konto or 'unbekannt'}', nicht vom Formular")
            )
            continue
        daten = parse_body(issue.get("body", ""))
        if daten is None or daten.get("art") not in ("neu", "aenderung"):
            uebersprungen.append(issue)
            continue
        # Erst pruefen, dann schreiben: ein Abbruch nach dem Speichern wuerde
        # beim naechsten Lauf Dubletten erzeugen (Excel geschrieben, Issue offen).
        grund = pruefe_eintrag(daten, bestand)
        if grund:
            abgelehnt.append((issue, grund))
            continue
        uebernehmen.append((issue, bereinigt(daten)))

    aenderungen = [d for _, d in uebernehmen if d.get("art") == "aenderung"]
    neue = [d for _, d in uebernehmen if d.get("art") != "aenderung"]

    # Ein einziger Schreibvorgang fuer den ganzen Lauf: er gelingt ganz oder gar
    # nicht. Bricht er ab, ist die Excel unveraendert und kein Issue geschlossen
    # – der Lauf laesst sich gefahrlos wiederholen.
    try:
        anzahl = in_excel_uebernehmen(XLSX, aenderungen, neue)
    except ZeileNichtGefunden as fehler:
        raise SystemExit(
            f"❌ Eine Aenderung liess sich nicht zuordnen: {fehler}\n\n"
            f"   Meist fehlt nur ein Zwischenschritt: Wurde {XLSX.name} von Hand\n"
            f"   geaendert, ohne danach build.bat zu starten, sucht das Programm\n"
            f"   noch nach dem alten Titel. Dann zuerst build.bat doppelklicken\n"
            f"   und vorschlaege.bat noch einmal starten.\n\n"
            f"   Steht der Skill gar nicht mehr in der Liste, wurde er inzwischen\n"
            f"   umbenannt oder entfernt: dann dem betroffenen Issue das Label\n"
            f"   `freigegeben` wegnehmen.\n\n"
            f"   Es wurde nichts veraendert – alle Vorschlaege bleiben freigegeben."
        )
    except ZeileMehrdeutig as fehler:
        raise SystemExit(
            f"❌ Eine Aenderung passt auf mehrere Zeilen: {fehler}\n\n"
            f"   Derselbe Titel steht in {XLSX.name} mehrfach in derselben Stufe\n"
            f"   und Kategorie. Welche Zeile gemeint ist, kann das Programm nicht\n"
            f"   wissen. Bitte oeffne die Datei, entferne oder benenne die\n"
            f"   doppelte Zeile um und starte vorschlaege.bat noch einmal.\n\n"
            f"   Es wurde nichts veraendert – alle Vorschlaege bleiben freigegeben."
        )

    nicht_geschlossen = []
    geschlossen = set()
    try:
        for issue, daten in uebernehmen:
            kennung = "~" if daten.get("art") == "aenderung" else "+"
            print(f"  {kennung} {daten['stufe']} / {daten['kategorie']}: {daten['titel']}")
            try:
                issue_schliessen(issue["number"])
                geschlossen.add(issue["number"])
            except subprocess.CalledProcessError:
                nicht_geschlossen.append(issue["number"])
    except BaseException:
        # Ab hier steht die Uebernahme bereits in der Excel. Schlaegt etwas
        # Unerwartetes durch (oder bricht jemand mit Strg+C ab), muss die
        # Warnung trotzdem erscheinen – sonst bleibt unbemerkt, dass Issues
        # offen sind und der naechste Lauf Dubletten erzeugt.
        warne_offene_issues(
            [i["number"] for i, _ in uebernehmen if i["number"] not in geschlossen]
        )
        raise

    if anzahl:
        wort = "Vorschlag" if anzahl == 1 else "Vorschläge"
        print(f"\n✅ {anzahl} {wort} in {XLSX.name} übernommen.")

    warne_offene_issues(nicht_geschlossen)

    for issue in uebersprungen:
        print(
            f'⚠ Issue #{issue["number"]} „{issue["title"]}" übersprungen '
            f'(kein lesbarer Vorschlagsblock oder andere Art) – bleibt offen.'
        )

    for issue, grund in abgelehnt:
        print(
            f'⚠ Issue #{issue["number"]} „{issue["title"]}" nicht übernommen: '
            f"{grund} – bleibt offen."
        )

    if uebersprungen or abgelehnt:
        print(
            f"\n   Was tun? Diese Issues auf github.com/{REPO}/issues anschauen.\n"
            f"   Taugt ein Vorschlag nichts, gib ihm das Label `abgelehnt` mit\n"
            f"   einer kurzen Begruendung als Kommentar und schliesse ihn. Soll er\n"
            f"   doch hinein, behebe den oben genannten Punkt (z. B. die Kategorie\n"
            f"   in der Excel wieder anlegen) und starte vorschlaege.bat noch\n"
            f"   einmal – das Label `freigegeben` kann stehen bleiben."
        )

    if anzahl:
        print("\nSkillsliste wird neu gebaut …\n")
        ergebnis = subprocess.run([sys.executable, str(BUILD)], check=False)
        if ergebnis.returncode != 0:
            print(
                "\n⚠ ACHTUNG: Der Neubau der Skillsliste ist gescheitert.\n"
                "   Was zu korrigieren ist, steht in der Meldung weiter oben.\n"
                f"   Die Vorschlaege stehen bereits in {XLSX.name} – nach der\n"
                "   Korrektur genuegt ein Doppelklick auf build.bat, ein erneuter\n"
                "   Lauf von vorschlaege.bat ist nicht noetig."
            )
            return
        print(
            "\nJetzt anschauen: docs/skillsliste.html\n"
            "Wenn es passt:  git add -A && git commit -m \"Neue Skills übernommen\" && git push"
        )


if __name__ == "__main__":
    main()
