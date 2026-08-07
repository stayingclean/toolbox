# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Skillsliste-Generator
=====================

Liest die Inhalte aus  skills_daten.xlsx  und erzeugt daraus die fertige
Webseite  docs/skillsliste.html  (auf Basis von  template.html).

Aufruf:  uv run build.py      (oder einfach build.bat doppelklicken)

Es wird KEIN Code in der HTML-Datei verändert – nur die Inhalte (Texte und
Emoji) werden aus der Excel-Datei eingesetzt. Das Layout/Design steckt
unverändert in template.html.
"""

import base64
import ipaddress
import json
import re
import sys
import unicodedata
from pathlib import Path
from urllib.parse import urlsplit

import openpyxl

# Konsole auf UTF-8 stellen, sonst scheitert das Drucken von Emoji/Umlauten.
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "skills_daten.xlsx"
TEMPLATE = ROOT / "template.html"
OUTPUT = ROOT / "docs" / "skillsliste.html"   # generierte Skillsliste-Seite
DATEN_JSON = ROOT / "docs" / "skills-daten.json"   # Datenstand für Formular + Worker
PLACEHOLDER = "var DATA = /*__BUILD_DATA__*/{};"
FAVICON_DIR = ROOT / "assets" / "favicons"
PLACEHOLDER_ICONS = "var ICONS = /*__BUILD_ICONS__*/{};"

TEMPLATE_VORSCHLAG = ROOT / "template-vorschlag.html"
OUTPUT_VORSCHLAG = ROOT / "docs" / "skill-vorschlagen.html"
PLACEHOLDER_VORSCHLAG = "var DATEN = /*__BUILD_DATA__*/{};"

# Anzeige-Name (Excel)  ->  interner Schlüssel (HTML/JS, darf sich NICHT ändern)
STUFE_KEY = {"Hoch": "hoch", "Mittel": "mittel", "Tief": "tief"}
STUFE_ORDER = ["hoch", "mittel", "tief"]

# Umlaute einsortieren wie im Wörterbuch (DIN 5007-1): ä zwischen a und b,
# nicht hinter z. Ohne diese Tabelle stünde „Duftöle" hinter „Duftzone", weil
# das ö im Zeichensatz nach dem z liegt. Das ß erledigt bereits casefold().
UMLAUTE = str.maketrans({"ä": "a", "ö": "o", "ü": "u"})


def sortier_schluessel(titel):
    """Wonach innerhalb einer Kategorie sortiert wird.

    Der zweite Teil ist der unveränderte Titel: Er entscheidet, wenn zwei
    Titel sich nur in Gross-/Kleinschreibung oder Umlauten unterscheiden.
    Damit hängt die Reihenfolge NICHT an der Zeilenfolge in der Excel — die
    Liste sieht nach einem Umsortieren der Mappe gleich aus wie vorher.
    """
    text = str(titel or "")
    return (text.casefold().translate(UMLAUTE), text)


# ── Bezugsquellen ────────────────────────────────────────────
# Dieselben Regeln stehen hier und in worker/validate.js (pruefeLinks).
# tools/vorschlaege_holen.py importiert sie von hier und kann darum nicht
# abweichen. Nur die JavaScript-Fassung muss von Hand nachgezogen werden —
# wird das vergessen, lässt der Worker einen Link durch, den der Build
# später ablehnt, und der Vorschlag steckt in der Excel fest.
LINK_SPALTEN = ["Link1", "Link2", "Link3"]
LINK_MAX_LAENGE = 300
TEXT_SPALTEN = ["Text1", "Text2", "Text3"]
# Adresse und Beschriftung gehoeren zusammen: sie werden gemeinsam gelesen,
# gemeinsam geprueft und rutschen beim Zusammenschieben gemeinsam auf.
LINK_PAARE = list(zip(LINK_SPALTEN, TEXT_SPALTEN))
TEXT_MAX_LAENGE = 30

# Linkverkürzer verbergen das Ziel vor der Freigabe – die Prüfung im Issue
# wäre wertlos – und ergeben als Knopfaufschrift nur "bit.ly" statt eines
# erkennbaren Händlers.
VERKUERZER = frozenset({
    "bit.ly", "tinyurl.com", "t.co", "goo.gl", "ow.ly", "is.gd",
    "buff.ly", "rb.gy", "cutt.ly", "shorturl.at", "s.id", "lnkd.in",
})


def _ist_ip(host: str) -> bool:
    try:
        ipaddress.ip_address(host)
        return True
    except ValueError:
        return False


def pruefe_link(roh: str) -> tuple[str | None, str | None]:
    """Prüft eine Bezugsquelle aus der Excel.

    Liefert (url, None) bei gültigem Link, sonst (None, Meldung). Der Build ist
    die letzte Schranke vor der Website: eine kaputte Adresse soll hier
    auffallen, nicht später einem Besucher beim Klicken.
    """
    url = str(roh or "").strip()
    if len(url) > LINK_MAX_LAENGE:
        return None, f"Link ist zu lang (höchstens {LINK_MAX_LAENGE} Zeichen)"
    # Spitze Klammern könnten in der erzeugten Skillsliste das <script>-Element
    # beenden. Die Ausgabecodierung in _render fängt das ab; dies ist die
    # zweite Schicht, wie schon bei den Textfeldern im Worker.
    if "<" in url or ">" in url:
        return None, "Link darf keine spitzen Klammern enthalten"
    try:
        teile = urlsplit(url)
    except ValueError:
        return None, "Link ist keine gültige Adresse"
    if teile.scheme != "https":
        return None, "Link muss mit https:// beginnen"
    if teile.username or teile.password:
        return None, "Link darf keine Benutzerangabe (@) enthalten"
    try:
        if teile.port is not None:
            return None, "Link darf keine Portnummer enthalten"
    except ValueError:
        return None, "Link hat eine ungültige Portnummer"
    host = (teile.hostname or "").strip(".")
    if _ist_ip(host):
        return None, "Link darf keine IP-Adresse sein"
    if "." not in host:
        return None, "Link hat keinen gültigen Hostnamen"
    if host.removeprefix("www.") in VERKUERZER:
        return None, "Linkverkürzer sind nicht erlaubt"
    return url, None


def pruefe_text(roh) -> tuple[str | None, str | None]:
    """Prüft die Beschriftung einer Bezugsquelle.

    Liefert (text, None) oder (None, Meldung). Kein `http`: die Domain steht
    nicht mehr im Knopf, darum fiele eine Beschriftung, die eine Adresse
    vortäuscht, kaum auf.
    """
    text = str(roh or "").strip()
    if len(text) > TEXT_MAX_LAENGE:
        return None, f"Beschriftung ist zu lang (höchstens {TEXT_MAX_LAENGE} Zeichen)"
    if "<" in text or ">" in text:
        return None, "Beschriftung darf keine spitzen Klammern enthalten"
    if "http" in text.lower():
        return None, "Beschriftung darf keine Adresse enthalten"
    return text, None


def gastgeber(url: str) -> str:
    """Hostname ohne führendes www. — der Schlüssel der Symboltabelle."""
    try:
        return (urlsplit(url).hostname or "").removeprefix("www.")
    except ValueError:
        return ""


def lade_favicons(data: dict) -> dict:
    """Symbol je Hostname als data:-URI.

    Fehlt eine Datei, kommt der Hostname schlicht nicht vor und der Knopf zeigt
    später den Hostnamen statt eines Bildes. Der Build darf daran NICHT
    scheitern: sonst legte eine neue Bezugsquelle bei einem unbekannten Händler
    die ganze Website lahm.
    """
    hosts = {
        gastgeber(link["u"])
        for stufe in data.values()
        for kat in stufe["kategorien"]
        for skill in kat["skills"]
        for link in skill["links"]
    }
    icons = {}
    for host in sorted(h for h in hosts if h):
        pfad = FAVICON_DIR / f"{host}.png"
        if not pfad.exists():
            continue
        icons[host] = "data:image/png;base64," + base64.b64encode(
            pfad.read_bytes()
        ).decode("ascii")
    return icons


class BuildError(Exception):
    """Sammelt eine oder mehrere verständliche Fehlermeldungen."""

    def __init__(self, messages):
        self.messages = messages if isinstance(messages, list) else [messages]
        super().__init__("\n".join(self.messages))


# ── Hilfsfunktionen ──────────────────────────────────────────
def clean(value) -> str:
    """Zellwert -> getrimmter String. Leere Zelle -> '' (nie 'None'/'nan')."""
    if value is None:
        return ""
    return str(value).strip()


def format_tip(tip: str) -> str:
    """Stellt der Birne 💡 voran (leerer Tipp bleibt leer).

    Idempotent: eine evtl. schon vorhandene führende Birne wird zuerst
    entfernt, damit nie '💡 💡 …' entsteht.
    """
    tip = tip.strip()
    if not tip:
        return ""
    if tip.startswith("💡"):
        tip = tip[1:].lstrip()
    return f"💡 {tip}"


def slug(label: str) -> str:
    """Erzeugt eine interne, stabile ID aus dem Kategorie-Label.

    Nur intern verwendet (aktiver Tab-Zustand im JS) – nie sichtbar.
    """
    text = label.lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    text = text.replace("ß", "ss")
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "kat"


def get_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise BuildError(
            f"Das Blatt '{name}' fehlt in skills_daten.xlsx. "
            f"Vorhandene Blätter: {', '.join(wb.sheetnames)}"
        )
    return wb[name]


def read_rows(ws, expected_header, optional_header=()):
    """Liest ein Blatt als Liste von dicts {Spaltenname: Wert}.

    Spalten aus `optional_header` werden gelesen, wenn sie vorhanden sind, und
    sonst als leerer String geliefert – so bricht eine ältere Excel nicht.
    Liefert zusätzlich die echte Excel-Zeilennummer für Fehlermeldungen.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BuildError(f"Das Blatt '{ws.title}' ist leer.")
    header = [clean(h) for h in rows[0]]
    missing = [h for h in expected_header if h not in header]
    if missing:
        raise BuildError(
            f"Im Blatt '{ws.title}' fehlen die Spalten: {', '.join(missing)}. "
            f"Bitte die Kopfzeile nicht umbenennen."
        )
    alle = list(expected_header) + list(optional_header)
    idx = {name: header.index(name) for name in alle if name in header}
    out = []
    for excel_row, raw in enumerate(rows[1:], start=2):
        record = {name: "" for name in alle}
        record.update(
            {name: clean(raw[i]) if i < len(raw) else "" for name, i in idx.items()}
        )
        if not any(record.values()):
            continue  # komplett leere Zeile überspringen
        record["_row"] = excel_row
        out.append(record)
    return out


# ── Daten lesen & prüfen ─────────────────────────────────────
def load_data():
    if not XLSX.exists():
        raise BuildError(f"Datei nicht gefunden: {XLSX.name}")

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    stufen_rows = read_rows(
        get_sheet(wb, "Stufen"),
        ["Stufe", "Bezeichnung", "Bereich", "Icon", "Intro", "Farbe", "Farbe2", "Hell"],
    )
    kat_rows = read_rows(get_sheet(wb, "Kategorien"), ["Stufe", "Kategorie", "Icon"])
    skill_rows = read_rows(
        get_sheet(wb, "Skills"),
        ["Stufe", "Kategorie", "Emoji", "Titel", "Beschreibung", "Tipp"],
        optional_header=["Von", "Ergaenzt"] + LINK_SPALTEN + TEXT_SPALTEN,
    )
    wb.close()

    errors = []

    def check_stufe(rec, blatt):
        if rec["Stufe"] not in STUFE_KEY:
            errors.append(
                f"Blatt '{blatt}', Zeile {rec['_row']}: Stufe '{rec['Stufe']}' "
                f"ist ungültig – erlaubt sind nur Hoch, Mittel oder Tief."
            )
            return None
        return STUFE_KEY[rec["Stufe"]]

    # Stufen-Metadaten
    stufen = {}
    for rec in stufen_rows:
        key = check_stufe(rec, "Stufen")
        if key:
            stufen[key] = rec
    fehlend = [s for s in STUFE_ORDER if s not in stufen]
    if fehlend:
        namen = ", ".join(k for k, v in STUFE_KEY.items() if v in fehlend)
        errors.append(f"Im Blatt 'Stufen' fehlen Zeilen für: {namen}.")

    # Kategorien (Reihenfolge je Stufe + Icon)
    kat_order = {s: [] for s in STUFE_ORDER}   # key -> [(label, icon)]
    kat_icon = {}                              # (key, label) -> icon
    for rec in kat_rows:
        key = check_stufe(rec, "Kategorien")
        if not key:
            continue
        label = rec["Kategorie"]
        if not label:
            errors.append(f"Blatt 'Kategorien', Zeile {rec['_row']}: Kategorie fehlt.")
            continue
        if (key, label) in kat_icon:
            errors.append(
                f"Blatt 'Kategorien', Zeile {rec['_row']}: Kategorie '{label}' "
                f"ist in Stufe '{rec['Stufe']}' doppelt."
            )
            continue
        kat_icon[(key, label)] = rec["Icon"]
        kat_order[key].append((label, rec["Icon"]))

    # Skills den Kategorien zuordnen (sortiert wird weiter unten)
    skills_by = {}  # (key, label) -> [skill-dict]
    for rec in skill_rows:
        key = check_stufe(rec, "Skills")
        if not key:
            continue
        label = rec["Kategorie"]
        for feld in ("Emoji", "Titel", "Beschreibung"):
            if not rec[feld]:
                errors.append(
                    f"Blatt 'Skills', Zeile {rec['_row']}: '{feld}' darf nicht leer sein."
                )
        if (key, label) not in kat_icon:
            errors.append(
                f"Blatt 'Skills', Zeile {rec['_row']}: Kategorie '{label}' "
                f"existiert in Stufe '{rec['Stufe']}' nicht "
                f"(bitte zuerst im Blatt 'Kategorien' anlegen)."
            )
            continue
        # Luecken werden zusammengeschoben: wer den ersten von zwei Links
        # entfernt, soll die uebrigen nicht von Hand aufruecken muessen. Die
        # Beschriftung wandert dabei mit ihrer Adresse mit.
        links = []
        for spalte, textspalte in LINK_PAARE:
            roh_url, roh_text = rec[spalte], rec[textspalte]
            if not roh_url:
                if roh_text:
                    errors.append(
                        f"Blatt 'Skills', Zeile {rec['_row']}, Spalte "
                        f"'{textspalte}': Beschriftung ohne Adresse in "
                        f"'{spalte}'."
                    )
                continue
            url, meldung = pruefe_link(roh_url)
            if meldung:
                errors.append(
                    f"Blatt 'Skills', Zeile {rec['_row']}, Spalte '{spalte}': "
                    f"{meldung}."
                )
                continue
            beschriftung, meldung = pruefe_text(roh_text)
            if meldung:
                errors.append(
                    f"Blatt 'Skills', Zeile {rec['_row']}, Spalte "
                    f"'{textspalte}': {meldung}."
                )
                continue
            if any(vorhanden["u"] == url for vorhanden in links):
                continue
            links.append({"u": url, "t": beschriftung})
        skills_by.setdefault((key, label), []).append(
            {
                "e": rec["Emoji"],
                "t": rec["Titel"],
                "b": rec["Beschreibung"],
                "tip": format_tip(rec["Tipp"]),
                "von": rec["Von"],
                "erg": rec["Ergaenzt"],
                "links": links,
            }
        )

    if errors:
        raise BuildError(errors)

    # Innerhalb jeder Kategorie alphabetisch nach Titel. Die Gruppierung nach
    # Stufe und Kategorie bleibt unberührt — die Reihenfolge der Kategorien
    # kommt weiterhin aus dem Blatt 'Kategorien'.
    # Zweck: Die Liste sieht immer gleich aus, egal in welcher Zeilenfolge die
    # Skills in der Excel stehen. Übernommene Vorschläge werden ans Ende der
    # Mappe angehängt und landeten sonst am Ende ihrer Kategorie.
    for liste in skills_by.values():
        liste.sort(key=lambda s: sortier_schluessel(s["t"]))

    # DATA-Struktur im erwarteten Schema bauen
    data = {}
    for key in STUFE_ORDER:
        meta = stufen[key]
        kategorien = []
        for label, icon in kat_order[key]:
            kategorien.append(
                {
                    "id": slug(label),
                    "label": label,
                    "icon": icon,
                    "skills": skills_by.get((key, label), []),
                }
            )
        data[key] = {
            "label": meta["Bezeichnung"],
            "bereich": meta["Bereich"],
            "farbe": meta["Farbe"],
            "farbe2": meta["Farbe2"],
            "hell": meta["Hell"],
            "icon": meta["Icon"],
            "intro": meta["Intro"],
            "kategorien": kategorien,
        }
    return data


# ── HTML schreiben ───────────────────────────────────────────
def _render(template_path, output_path, ersetzungen):
    """Setzt mehrere Platzhalter in eine Vorlage ein.

    `ersetzungen` ist eine Liste von (Platzhalter, Muster, Daten).
    """
    if not template_path.exists():
        raise BuildError(f"Vorlage nicht gefunden: {template_path.name}")
    html = template_path.read_bytes().decode("utf-8-sig")  # evtl. BOM entfernen
    for platzhalter, muster, daten in ersetzungen:
        if platzhalter not in html:
            raise BuildError(
                f"Platzhalter nicht in {template_path.name} gefunden. "
                f"Die Vorlage muss '{platzhalter}' enthalten."
            )
        payload = json.dumps(daten, ensure_ascii=False, separators=(", ", ": "))
        # Ausgabecodierung fuer den <script>-Block: sonst beendet ein </script>
        # im Freitext das Skriptelement und der Rest wird als Markup ausgefuehrt.
        payload = (
            payload.replace("<", "\\u003c")
            .replace(">", "\\u003e")
            .replace("&", "\\u0026")
        )
        html = html.replace(platzhalter, muster % payload, 1)
    # gleiche Datei-Konvention wie das Original: UTF-8 mit BOM
    output_path.write_bytes(b"\xef\xbb\xbf" + html.encode("utf-8"))


def render(data: dict, icons: dict):
    _render(
        TEMPLATE,
        OUTPUT,
        [
            (PLACEHOLDER, "var DATA = %s;", data),
            (PLACEHOLDER_ICONS, "var ICONS = %s;", icons),
        ],
    )


def render_vorschlag(data: dict):
    _render(
        TEMPLATE_VORSCHLAG,
        OUTPUT_VORSCHLAG,
        [(PLACEHOLDER_VORSCHLAG, "var DATEN = %s;", data)],
    )


def write_daten_json(data: dict):
    """Schreibt den Datenstand für die Vorschlagsseite und den Worker.

    Bewusst OHNE BOM: JSON.parse im Worker stolpert sonst über das erste Zeichen.
    Zeilenenden fest auf Unix-Art und ein Zeilenumbruch am Schluss – sonst
    erzeugt ein Build auf einem anderen Betriebssystem einen Scheinunterschied
    über die ganze Datei.
    """
    DATEN_JSON.write_text(
        json.dumps(data, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def main():
    try:
        data = load_data()
        render(data, lade_favicons(data))
        render_vorschlag(data)
        write_daten_json(data)
    except BuildError as exc:
        print("\n❌ Build abgebrochen – bitte folgendes in skills_daten.xlsx korrigieren:\n")
        for msg in exc.messages:
            print(f"   • {msg}")
        print()
        sys.exit(1)

    total = sum(len(k["skills"]) for d in data.values() for k in d["kategorien"])
    print(f"✅ {OUTPUT.relative_to(ROOT).as_posix()} wurde neu erstellt.")
    print(f"✅ {OUTPUT_VORSCHLAG.relative_to(ROOT).as_posix()} wurde neu erstellt.")
    print(f"✅ {DATEN_JSON.relative_to(ROOT).as_posix()} wurde neu erstellt.")
    print(f"   Stufen: {len(data)} | Kategorien: "
          f"{sum(len(d['kategorien']) for d in data.values())} | Skills: {total}")


if __name__ == "__main__":
    main()
